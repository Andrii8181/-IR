# sad_journal_trial.py — Польовий журнал, генератор плану досліду
# -*- coding: utf-8 -*-
from sad_common import *
from sad_homogeneous import (HPPlant, HP_ROLE_RECORDED, HP_ROLE_GUARD_EDGE,
    HP_ROLE_GUARD_REP, HP_ROLE_DEAD, HP_ROLE_POLLINIZER, HP_ROLE_UNASSIGNED,
    HP_ROLE_EXTRA, HP_ROLE_EXCLUDED_CV, HP_ROLE_COLORS, HP_ROLE_LABELS)
from sad_repeated import MixedRepeatedWindow


# ═══════════════════════════════════════════════════════════════
# СПІЛЬНІ ФУНКЦІЇ «ВІДКРИТИ ПОКАЗНИК» — використовуються з будь-якого
# модуля аналізу (ANOVA, регресія, кореляція, PCA, кластерний) для
# імпорту даних, збережених у польовому журналі, без ручного
# копіювання-вставляння.
# ═══════════════════════════════════════════════════════════════

def load_journal_file(parent):
    """Відкриває діалог вибору файлу журналу (.sadp, type=field_journal).
    Повертає dict {"plants","factor_defs","variant_names","records"} або
    None, якщо скасовано чи файл не підходить."""
    path = filedialog.askopenfilename(
        parent=parent, filetypes=[("SAD журнал","*.sadp"),("JSON","*.json")],
        title="Відкрити журнал обліків")
    if not path: return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            d = json.load(f)
    except Exception as ex:
        messagebox.showerror("Помилка відкриття", str(ex)); return None
    if d.get("type") != "field_journal":
        messagebox.showwarning("Не той тип файлу",
            "Показники зберігаються в ЖУРНАЛІ обліків, а не в самій схемі без "
            "внесених даних. Відкрийте файл, збережений через «💾 Зберегти "
            "журнал» у польовому журналі обліків."); return None

    plants = []
    for pd in d.get("plants", []):
        p = HPPlant(pd["row"], pd["position"], pd.get("value"), pd.get("status","ok"))
        p.role = pd.get("role", HP_ROLE_UNASSIGNED)
        p.variant = pd.get("variant")
        p.replication = pd.get("replication")
        p.factors = pd.get("factors", {})
        plants.append(p)

    records = d.get("records", {})
    if not records:
        messagebox.showwarning("Немає обліків",
            "У цьому журналі ще немає жодного внесеного показника."); return None

    return {"plants": plants, "factor_defs": d.get("factor_defs", []),
            "variant_names": d.get("variant_names", []), "records": records,
            "path": path}


def aggregate_journal_data(plants, factor_defs, variant_names, records, record_names):
    """Зводить журнал у таблицю: один рядок на КОЖНУ повторність — значення
    облікових рослин У МЕЖАХ однієї повторності усереднюються в одне число
    (уникає псевдоповторності). Повертає (factor_cols, rows)."""
    is_multi = bool(factor_defs) and any(p.factors for p in plants)
    groups = {}
    for p in plants:
        if p.role != HP_ROLE_RECORDED: continue
        if is_multi:
            key = (tuple(sorted(p.factors.items())), p.replication)
        else:
            key = (p.variant, p.replication)
        groups.setdefault(key, []).append(p)

    factor_cols = [fdef["name"] for fdef in factor_defs] if is_multi else ["Варіант"]
    rows = []
    for key in sorted(groups.keys(), key=lambda k: (str(k[0]), k[1] or 0)):
        plist = groups[key]
        row = {"replication": key[1], "n_subsamples": len(plist)}
        if is_multi:
            factor_levels = dict(key[0])
            for fdef in factor_defs:
                fname = fdef["name"]; lv = factor_levels.get(fname)
                lvl_name = (fdef["levels"][lv-1] if lv and 1 <= lv <= len(fdef["levels"])
                           else f"Рівень {lv}")
                row[fname] = lvl_name
        else:
            v = key[0]
            vname = (variant_names[v-1] if variant_names and v and
                     1 <= v <= len(variant_names) else f"В{v}")
            row["Варіант"] = vname
        for rn in record_names:
            vals_dict = records.get(rn, {}).get("values", {})
            vals = []
            for p in plist:
                vkey = f"{p.row}:{p.position}"
                raw = vals_dict.get(vkey)
                if raw is None or str(raw).strip() == "": continue
                try: vals.append(float(str(raw).replace(",",".")))
                except ValueError: pass
            row[rn] = round(sum(vals)/len(vals), 4) if vals else None
        rows.append(row)
    return factor_cols, rows


def aggregate_to_variant_level(factor_cols, rows, record_names):
    """Друге усереднення — рядки (варіант×повторність), вже усереднені в
    межах повторності, усереднюються ЩЕ РАЗ по повторностях: одне число
    на варіант для кожного показника. Стандартний підхід для порівняння
    варіантів/сортів за кількома показниками одночасно (регресія,
    кореляція, PCA, кластерний аналіз)."""
    groups = {}
    for r in rows:
        key = tuple(r[fc] for fc in factor_cols)
        groups.setdefault(key, []).append(r)
    out = []
    for key in sorted(groups.keys()):
        plist = groups[key]
        row = dict(zip(factor_cols, key))
        for rn in record_names:
            vals = [r[rn] for r in plist if r[rn] is not None]
            row[rn] = round(sum(vals)/len(vals), 4) if vals else None
        out.append(row)
    return out


def pick_indicators_dialog(parent, records, multi_select=False, n_required=None):
    """Показує список показників журналу для вибору. multi_select=False —
    один показник (радіо), True — довільна кількість чекбоксами.
    n_required — якщо задано (напр. 2 для регресії), вимагає рівно
    стільки позначених. Повертає list обраних назв, або None якщо
    скасовано."""
    names = list(records.keys())
    result = {"value": None}
    dlg = tk.Toplevel(parent); dlg.title("Оберіть показник" + ("и" if multi_select else ""))
    dlg.resizable(False, False); set_icon(dlg); dlg.grab_set()
    rf = ("Times New Roman",11)
    frm = tk.Frame(dlg, padx=18, pady=14); frm.pack()
    tk.Label(frm, text="Оберіть показник" + (f" (рівно {n_required})" if n_required else
             " (один чи кілька)" if multi_select else "") + ":",
             font=("Times New Roman",11,"bold")).pack(anchor="w", pady=(0,8))

    if multi_select:
        vars_ = {nm: tk.BooleanVar(value=False) for nm in names}
        for nm in names:
            unit = records[nm].get("unit","")
            tk.Checkbutton(frm, text=f"{nm}" + (f" ({unit})" if unit else ""),
                          variable=vars_[nm], font=rf).pack(anchor="w")
    else:
        sel = tk.StringVar(value=names[0] if names else "")
        for nm in names:
            unit = records[nm].get("unit","")
            tk.Radiobutton(frm, text=f"{nm}" + (f" ({unit})" if unit else ""),
                          variable=sel, value=nm, font=rf).pack(anchor="w")

    def _ok():
        if multi_select:
            chosen = [nm for nm in names if vars_[nm].get()]
            if n_required and len(chosen) != n_required:
                messagebox.showwarning("",
                    f"Оберіть рівно {n_required} показники.", parent=dlg); return
            if not chosen:
                messagebox.showwarning("", "Оберіть хоча б один показник.", parent=dlg); return
            result["value"] = chosen
        else:
            if not sel.get():
                messagebox.showwarning("", "Оберіть показник.", parent=dlg); return
            result["value"] = [sel.get()]
        dlg.destroy()

    bf = tk.Frame(frm); bf.pack(pady=(12,0))
    tk.Button(bf, text="Обрати", bg="#1a6b1a", fg="white", font=rf,
              command=_ok).pack(side=tk.LEFT, padx=4)
    tk.Button(bf, text="Скасувати", font=rf, command=dlg.destroy).pack(side=tk.LEFT)
    center_win(dlg)
    dlg.wait_window()
    return result["value"]


def open_indicator_for_anova(parent):
    """Для ANOVA: варіант(и) × повторність. Повертає (factor_cols, rows)
    або None. rows містять "replication" і значення показника —
    придатні для прямого широкоформатного розкладання в таблицю ANOVA."""
    data = load_journal_file(parent)
    if data is None: return None
    chosen = pick_indicators_dialog(parent, data["records"], multi_select=False)
    if not chosen: return None
    factor_cols, rows = aggregate_journal_data(
        data["plants"], data["factor_defs"], data["variant_names"], data["records"], chosen)
    return factor_cols, rows, chosen[0]


def open_indicators_for_variant_analysis(parent, multi_select=True, n_required=None):
    """Для регресії/кореляції/PCA/кластерного: варіант-рівень (подвійне
    усереднення). Повертає (factor_cols, rows, record_names) або None.
    rows — по одному запису на варіант, значення — вже усереднені і в
    межах повторності, і по повторностях."""
    data = load_journal_file(parent)
    if data is None: return None
    chosen = pick_indicators_dialog(parent, data["records"],
                                    multi_select=multi_select, n_required=n_required)
    if not chosen: return None
    factor_cols, rows = aggregate_journal_data(
        data["plants"], data["factor_defs"], data["variant_names"], data["records"], chosen)
    variant_rows = aggregate_to_variant_level(factor_cols, rows, chosen)
    return factor_cols, variant_rows, chosen


class FieldJournalWindow:
    """
    Окремий модуль від «Планування досліду за однорідністю» (навмисно —
    для простоти й зрозумілості кожен інструмент робить одну річ).

    Завантажує ЗБЕРЕЖЕНУ СХЕМУ (файл .sadp, type="homogeneous_plot_scheme")
    або вже існуючий журнал (type="field_journal") і дозволяє:
      • вписувати виміряні значення показника прямо в клітинки схеми —
        на екрані комп'ютера, у тому самому фізичному розташуванні, що
        й у саду;
      • вести кілька РІЗНИХ показників («Урожайність 2024», «Діаметр
        штамбу 2023» тощо) по ОДНІЙ і тій самій фізичній схемі —
        кожен зберігається як окремий «облік» (record) у тому самому
        файлі журналу;
      • роздрукувати порожній бланк для внесення даних у польових
        умовах олівцем/ручкою.
    """

    HELP_TEXT = """
ПОЛЬОВИЙ ЖУРНАЛ ОБЛІКІВ — ІНСТРУКЦІЯ
═════════════════════════════════════

ДЛЯ ЧОГО ЦЕЙ МОДУЛЬ?
  Модуль «Планування досліду за однорідністю» створює схему — яка саме
  рослина до якого варіанту й повторення належить. Але сама схема не
  містить майбутніх результатів обліків (урожайність, вміст цукру
  тощо) — для цього і є цей журнал.

КРОК 1. ВІДКРИЙТЕ СХЕМУ АБО ЖУРНАЛ
  «📂 Відкрити» — оберіть файл .sadp:
    • якщо це щойно збережена СХЕМА — журнал створюється з нуля на її
      основі (ще без жодного обліку);
    • якщо це вже існуючий ЖУРНАЛ — відкриються всі попередні обліки,
      які в ньому вже збережені.

КРОК 2. СТВОРІТЬ ОБЛІК (ПОКАЗНИК)
  «➕ Новий облік» — вкажіть назву показника й одиницю виміру
  (наприклад, «Урожайність 2024», кг). Можна створити скільки завгодно
  обліків по одній і тій самій схемі — для різних років чи показників.
  Перемикайтесь між ними через випадаючий список зверху.

КРОК 3. ВНОСЬТЕ ЗНАЧЕННЯ
  Зелені клітинки — облікові рослини поточного варіанту/повторення,
  саме туди вписуються виміряні значення. Сірі/жовті/помаранчеві —
  захисні/виключені рослини (не редагуються, показані лише для
  орієнтації в саду). Дані зберігаються автоматично в пам'яті —
  не забудьте «💾 Зберегти журнал» наприкінці роботи.

ДРУК
  «🖨 Друкувати бланк» — та сама схема, але порожня, з місцем для
  запису значень від руки в полі (як в модулі планування).
"""

    def __init__(self, parent, gs=None):
        self._parent = parent
        self.win = tk.Toplevel(parent)
        self.win.title("Польовий журнал обліків")
        self.win.geometry("1400x820"); set_icon(self.win)
        self.gs = dict(gs) if gs else {}
        self.plants = []           # список HPPlant з завантаженої схеми
        self.cfg = {}
        self.variant_names = []
        self.factor_defs = []      # [{"name":...,"levels":[...]}, ...] — для багатофакторних схем
        self.records = {}          # {"Урожайність 2024": {"unit":..., "values": {"r:p": val}}}
        self.current_record = None
        self._journal_path = None
        self._entry_widgets = {}   # (row,position) -> tk.Entry (лише для облікових)
        self._build()

    # ─────────────────────────────────────────────────────
    def _build(self):
        rf = ("Times New Roman", 11)
        top = tk.Frame(self.win, padx=8, pady=6); top.pack(fill=tk.X)
        tk.Button(top, text="📂 Відкрити схему / журнал", bg="#1a4b8c", fg="white",
                  font=rf, command=self._open_file).pack(side=tk.LEFT, padx=4)
        tk.Button(top, text="💾 Зберегти журнал", bg="#1a6b1a", fg="white",
                  font=rf, command=self._save_journal).pack(side=tk.LEFT, padx=4)
        tk.Button(top, text="➕ Новий облік", font=rf,
                  command=self._new_record).pack(side=tk.LEFT, padx=(16,4))

        tk.Label(top, text="Поточний облік:", font=rf).pack(side=tk.LEFT, padx=(16,2))
        self._record_var = tk.StringVar(value="")
        self._record_cb = ttk.Combobox(top, textvariable=self._record_var,
                                       state="readonly", width=28, values=[])
        self._record_cb.pack(side=tk.LEFT, padx=2)
        self._record_cb.bind("<<ComboboxSelected>>",
                             lambda e: self._switch_record(self._record_var.get()))

        tk.Button(top, text="🖨 Друкувати бланк", font=rf,
                  command=self._print_blank_form).pack(side=tk.LEFT, padx=(16,4))
        tk.Button(top, text="📊 Звести для аналізу", bg="#8c1a1a", fg="white", font=rf,
                  command=self._open_aggregate_dialog).pack(side=tk.LEFT, padx=4)
        tk.Button(top, text="📚 Довідка", bg="#1a4b8c", fg="white", font=rf,
                  command=self._show_help).pack(side=tk.LEFT, padx=4)

        self._info_lbl = tk.Label(self.win,
                text="Немає завантаженої схеми. Натисніть «📂 Відкрити схему / журнал», "
                     "щоб почати.",
                font=("Times New Roman",11), fg="#888", anchor="w")
        self._info_lbl.pack(fill=tk.X, padx=8, pady=(0,4))

        # ── Легенда (та сама, що й у схемі) ───────────────
        self._legend_f = tk.Frame(self.win, bg="#f7f7f7", padx=8, pady=6)
        self._legend_f.pack(fill=tk.X)

        # ── Прокручувана область сітки ─────────────────────
        tbl_area = tk.Frame(self.win); tbl_area.pack(fill=tk.BOTH, expand=True, padx=8, pady=(2,4))
        self._canvas = tk.Canvas(tbl_area)
        self._canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb_v = ttk.Scrollbar(tbl_area, orient="vertical", command=self._canvas.yview)
        sb_v.pack(side=tk.RIGHT, fill=tk.Y)
        sb_h = ttk.Scrollbar(self.win, orient="horizontal", command=self._canvas.xview)
        sb_h.pack(fill=tk.X, padx=8)
        self._canvas.configure(yscrollcommand=sb_v.set, xscrollcommand=sb_h.set)
        self.inner = tk.Frame(self._canvas)
        self._canvas.create_window((0,0), window=self.inner, anchor="nw")
        self.inner.bind("<Configure>",
                        lambda e: self._canvas.config(scrollregion=self._canvas.bbox("all")))
        self.win.bind("<MouseWheel>",
                      lambda e: self._canvas.yview_scroll(int(-1*(e.delta/120)),"units"))

    def _show_help(self):
        win = tk.Toplevel(self.win); win.title("Довідка — Польовий журнал")
        win.geometry("700x600"); set_icon(win)
        txt = tk.Text(win, wrap="word", font=("Times New Roman",11), padx=10, pady=10)
        txt.pack(fill=tk.BOTH, expand=True)
        txt.insert("1.0", self.HELP_TEXT.strip()); txt.configure(state="disabled")
        tk.Button(win, text="Закрити", command=win.destroy,
                  font=("Times New Roman",11)).pack(pady=6)

    # ── Відкриття схеми/журналу ───────────────────────────
    def _open_file(self):
        path = filedialog.askopenfilename(
            parent=self.win, filetypes=[("SAD файл","*.sadp"),("JSON","*.json")],
            title="Відкрити схему або журнал")
        if not path: return
        try:
            with open(path, "r", encoding="utf-8") as f:
                d = json.load(f)
        except Exception as ex:
            messagebox.showerror("Помилка відкриття", str(ex)); return

        ftype = d.get("type")
        if ftype not in ("homogeneous_plot_scheme", "multi_factor_scheme", "field_journal"):
            messagebox.showwarning("Не той тип файлу",
                "Цей файл не є ані схемою досліду, ані журналом обліків.\n"
                "Оберіть файл, збережений через «💾 Зберегти схема» в модулі "
                "планування чи конструкторі схеми, або раніше збережений журнал."); return

        plants = []
        for pd in d.get("plants", []):
            p = HPPlant(pd["row"], pd["position"], pd.get("value"), pd.get("status","ok"))
            p.role = pd.get("role", HP_ROLE_UNASSIGNED)
            p.plot_id = pd.get("plot_id")
            p.variant = pd.get("variant")
            p.replication = pd.get("replication")
            p.factors = pd.get("factors", {})
            plants.append(p)
        self.plants = plants
        self.cfg = d.get("cfg", {})
        self.variant_names = d.get("variant_names", [])
        self.factor_defs = d.get("factor_defs", [])   # [{"name":"Обробка ґрунту","levels":["Оранка","Без оранки"]}, ...]
        self.records = d.get("records", {}) if ftype == "field_journal" else {}
        self._journal_path = path if ftype == "field_journal" else None

        self._record_cb.configure(values=list(self.records.keys()))
        if self.records:
            first = list(self.records.keys())[0]
            self._record_var.set(first)
            self.current_record = first
        else:
            self._record_var.set("")
            self.current_record = None

        n_rec = len(self.records)
        extra_warn = ("   |   ⚠ Спочатку натисніть «➕ Новий облік»" if n_rec == 0 else "")
        self._info_lbl.configure(
            text=f"Завантажено: {os.path.basename(path)}   |   "
                 f"Показник схеми: {self.cfg.get('trait_name','—')}   |   "
                 f"Облікових рослин: {sum(1 for p in self.plants if p.role == HP_ROLE_RECORDED)}   |   "
                 f"Обліків у журналі: {n_rec}{extra_warn}",
            fg=("#c62828" if n_rec == 0 else "#1a6b1a"))
        self._build_legend()
        self._build_grid()

    # ── Керування обліками (показниками) ──────────────────
    def _new_record(self):
        if not self.plants:
            messagebox.showwarning("", "Спочатку відкрийте схему або журнал."); return
        dlg = tk.Toplevel(self.win); dlg.title("Новий облік")
        dlg.resizable(False, False); set_icon(dlg); dlg.grab_set()
        frm = tk.Frame(dlg, padx=16, pady=14); frm.pack()
        tk.Label(frm, text="Назва показника (напр. «Урожайність 2024»):",
                 font=("Times New Roman",11)).grid(row=0, column=0, sticky="w", pady=4)
        name_v = tk.StringVar()
        tk.Entry(frm, textvariable=name_v, width=30, font=("Times New Roman",11)
                 ).grid(row=0, column=1, sticky="w", padx=8)
        tk.Label(frm, text="Одиниця виміру:", font=("Times New Roman",11)
                 ).grid(row=1, column=0, sticky="w", pady=4)
        unit_v = tk.StringVar()
        tk.Entry(frm, textvariable=unit_v, width=15, font=("Times New Roman",11)
                 ).grid(row=1, column=1, sticky="w", padx=8)

        def _create():
            nm = name_v.get().strip()
            if not nm:
                messagebox.showwarning("", "Вкажіть назву показника.", parent=dlg); return
            if nm in self.records:
                messagebox.showwarning("", "Такий облік вже існує.", parent=dlg); return
            self.records[nm] = {"unit": unit_v.get().strip(), "values": {}}
            self._record_cb.configure(values=list(self.records.keys()))
            self._record_var.set(nm)
            self.current_record = nm
            dlg.destroy()
            self._build_grid()
        bf = tk.Frame(frm); bf.grid(row=2, column=0, columnspan=2, pady=(12,0))
        tk.Button(bf, text="Створити", bg="#1a6b1a", fg="white",
                  font=("Times New Roman",11), command=_create).pack(side=tk.LEFT, padx=4)
        tk.Button(bf, text="Скасувати", font=("Times New Roman",11),
                  command=dlg.destroy).pack(side=tk.LEFT)
        center_win(dlg)

    def _switch_record(self, name):
        if not name or name not in self.records: return
        self._save_current_values_from_widgets()
        self.current_record = name
        self._build_grid()

    def _save_current_values_from_widgets(self):
        if self.current_record is None: return
        vals = self.records[self.current_record]["values"]
        for (r,p), e in self._entry_widgets.items():
            txt = e.get().strip()
            key = f"{r}:{p}"
            if txt: vals[key] = txt
            elif key in vals: del vals[key]

    # ── Легенда (компактна, та сама палітра що й у схемі) ─
    def _build_legend(self):
        for w in self._legend_f.winfo_children(): w.destroy()
        row1 = tk.Frame(self._legend_f, bg="#f7f7f7"); row1.pack(fill=tk.X)
        for role, color in HP_ROLE_COLORS.items():
            sw = tk.Frame(row1, bg=color, width=14, height=14, relief=tk.RIDGE, bd=1)
            sw.pack(side=tk.LEFT, padx=(0,4), pady=2); sw.pack_propagate(False)
            tk.Label(row1, text=HP_ROLE_LABELS[role], bg="#f7f7f7",
                     font=("Times New Roman",9)).pack(side=tk.LEFT, padx=(0,12))
        if self.variant_names:
            names_txt = "   •   ".join(f"В{i+1}={nm}" for i, nm in enumerate(self.variant_names))
            tk.Label(self._legend_f, text="Варіанти: " + names_txt,
                     bg="#f7f7f7", fg="#1a4b8c", font=("Times New Roman",9,"bold"),
                     anchor="w").pack(fill=tk.X, pady=(2,0))

    # ── Основна сітка внесення даних ───────────────────────
    def _build_grid(self):
        for w in self.inner.winfo_children(): w.destroy()
        self._entry_widgets = {}
        if not self.plants: return

        by_row = {}
        for p in self.plants:
            by_row.setdefault(p.row, {})[p.position] = p
        rows = sorted(by_row.keys())
        max_pos = max((p.position for p in self.plants), default=1)

        vals = self.records.get(self.current_record, {}).get("values", {}) \
               if self.current_record else {}
        no_record = self.current_record is None

        if no_record:
            tk.Label(self.inner,
                     text="⚠ Спочатку створіть облік («➕ Новий облік») — доки його "
                          "немає, клітинки недоступні для вводу, щоб уникнути втрати "
                          "введених значень.",
                     font=("Times New Roman",10,"bold"), fg="#c62828", bg="#fff3e0",
                     anchor="w", padx=6, pady=4
                     ).grid(row=0, column=0, columnspan=max(2,max_pos+1), sticky="ew")
            header_row = 1
        else:
            header_row = 0

        tk.Label(self.inner, text="Ряд \\ Поз.", width=9, relief=tk.RIDGE,
                 bg="#444444", fg="white", font=("Times New Roman",10,"bold")
                 ).grid(row=header_row, column=0, padx=1, pady=1, sticky="nsew")
        for j in range(1, max_pos+1):
            tk.Label(self.inner, text=str(j), width=6, relief=tk.RIDGE,
                     bg="#1a4b8c", fg="white", font=("Times New Roman",9,"bold")
                     ).grid(row=header_row, column=j, padx=1, pady=1, sticky="nsew")

        for ri, row_num in enumerate(rows):
            grid_row = header_row + 1 + ri
            tk.Label(self.inner, text=f"Ряд {row_num}", width=9, relief=tk.RIDGE,
                     bg="#444444", fg="white", font=("Times New Roman",9,"bold")
                     ).grid(row=grid_row, column=0, padx=1, pady=1, sticky="nsew")
            for pos in range(1, max_pos+1):
                p = by_row.get(row_num, {}).get(pos)
                if p is None:
                    continue
                if p.role == HP_ROLE_RECORDED:
                    key = f"{row_num}:{pos}"
                    e = tk.Entry(self.inner, width=6, font=("Times New Roman",9),
                                bg=("#eeeeee" if no_record else "#e8f5e9"),
                                justify="center",
                                state=(tk.DISABLED if no_record else tk.NORMAL))
                    e.grid(row=grid_row, column=pos, padx=1, pady=1)
                    if not no_record:
                        if key in vals: e.insert(0, str(vals[key]))
                        e.bind("<FocusOut>", lambda ev, r=row_num, po=pos: self._on_cell_edit(r, po))
                        self._entry_widgets[(row_num,pos)] = e
                else:
                    label = {HP_ROLE_GUARD_EDGE:"К", HP_ROLE_GUARD_REP:"П",
                             HP_ROLE_DEAD:"-", HP_ROLE_POLLINIZER:"+",
                             HP_ROLE_EXTRA:"×", HP_ROLE_EXCLUDED_CV:"✕"}.get(p.role, "")
                    color = HP_ROLE_COLORS.get(p.role, "#eeeeee")
                    tk.Label(self.inner, text=label, width=6, relief=tk.RIDGE,
                             bg=color, font=("Times New Roman",9)
                             ).grid(row=grid_row, column=pos, padx=1, pady=1)
        if not no_record:
            _bind_nav(self._grid_as_2d(rows, max_pos, by_row), self.win)

    def _grid_as_2d(self, rows, max_pos, by_row):
        """Формує 2D-масив лише з редагованих (облікових) клітинок —
        для навігації Enter/стрілками (пропускаючи нередаговані)."""
        out = []
        for row_num in rows:
            line = []
            for pos in range(1, max_pos+1):
                p = by_row.get(row_num, {}).get(pos)
                if p is not None and p.role == HP_ROLE_RECORDED:
                    e = self._entry_widgets.get((row_num,pos))
                    if e is not None: line.append(e)
            if line: out.append(line)
        return out

    def _on_cell_edit(self, row, pos):
        if self.current_record is None: return
        e = self._entry_widgets.get((row,pos))
        if e is None: return
        vals = self.records[self.current_record]["values"]
        key = f"{row}:{pos}"
        txt = e.get().strip()
        if txt: vals[key] = txt
        elif key in vals: del vals[key]

    # ── Збереження журналу ──────────────────────────────────
    def _save_journal(self):
        if not self.plants:
            messagebox.showwarning("", "Немає завантаженої схеми — нічого зберігати."); return
        self._save_current_values_from_widgets()
        plants_data = []
        for p in self.plants:
            plants_data.append({
                "row": p.row, "position": p.position, "value": p.value,
                "status": p.status, "role": p.role, "plot_id": p.plot_id,
                "variant": p.variant, "replication": p.replication,
                "factors": p.factors,
            })
        d = {
            "type": "field_journal", "version": APP_VER,
            "cfg": self.cfg, "variant_names": self.variant_names,
            "factor_defs": self.factor_defs,
            "plants": plants_data, "records": self.records,
        }
        default_name = os.path.basename(self._journal_path) if self._journal_path else \
            "журнал_" + (self.cfg.get("trait_name","досліду").replace(" ","_")) + ".sadp"
        path = filedialog.asksaveasfilename(
            parent=self.win, defaultextension=".sadp", initialfile=default_name,
            filetypes=[("SAD журнал","*.sadp"),("JSON","*.json")],
            title="Зберегти журнал")
        if not path: return
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(d, f, ensure_ascii=False, indent=2)
            self._journal_path = path
            messagebox.showinfo("Збережено", f"Журнал збережено:\n{path}")
        except Exception as ex:
            messagebox.showerror("Помилка збереження", str(ex))

    # ── Друк порожнього бланку ──────────────────────────────
    def _print_blank_form(self):
        if not self.plants:
            messagebox.showwarning("", "Спочатку відкрийте схему або журнал."); return
        if not HAS_MPL:
            messagebox.showwarning("", "Для друку потрібен matplotlib."); return

        by_row = {}
        for p in self.plants:
            by_row.setdefault(p.row, {})[p.position] = p
        rows = sorted(by_row.keys())

        POS_PER_PAGE = 14; ROWS_PER_PAGE = 4
        segments = []
        for row_num in rows:
            positions = sorted(by_row[row_num].keys())
            for i in range(0, len(positions), POS_PER_PAGE):
                segments.append((row_num, positions[i:i+POS_PER_PAGE]))
        pages = [segments[i:i+ROWS_PER_PAGE] for i in range(0, len(segments), ROWS_PER_PAGE)] \
                or [[]]

        page_idx = {"i": 0}
        win = tk.Toplevel(self.win); win.title("Друк бланку журналу")
        win.geometry("1100x750"); set_icon(win)
        tb = tk.Frame(win, padx=6, pady=5); tb.pack(fill=tk.X)
        lbl = tk.Label(tb, text="", font=("Times New Roman",11,"bold"))
        lbl.pack(side=tk.LEFT, padx=8)
        holder = tk.Frame(win); holder.pack(fill=tk.BOTH, expand=True)
        state = {"fig": None}

        def _render():
            for w in holder.winfo_children(): w.destroy()
            segs = pages[page_idx["i"]]
            n_seg = max(1, len(segs))
            max_len = max((len(pl) for _, pl in segs), default=1)
            fig = Figure(figsize=(max(9, max_len*0.85), max(4.5, n_seg*2.3+1.2)), dpi=100)
            ax = fig.add_subplot(111)
            ROW_H = 2.1
            for si, (row_num, pos_list) in enumerate(segs):
                y_top = n_seg*ROW_H - si*ROW_H
                ax.text(-0.7, y_top-1.0, f"Ряд {row_num}\n(поз. {pos_list[0]}-{pos_list[-1]})",
                        ha="right", va="center", fontsize=9, fontfamily="Times New Roman",
                        fontweight="bold")
                for ci, pos in enumerate(pos_list):
                    p = by_row.get(row_num, {}).get(pos)
                    if p is None: continue
                    x = ci
                    ax.text(x+0.5, y_top+0.15, str(pos), ha="center", va="center",
                            fontsize=8, fontfamily="Times New Roman", color="#555")
                    if p.role == HP_ROLE_RECORDED:
                        top_r = matplotlib.patches.Rectangle(
                            (x+0.03, y_top-0.72), 0.94, 0.55,
                            facecolor="#EAF2FB", edgecolor="#333", linewidth=1.0)
                        ax.add_patch(top_r)
                        ax.text(x+0.5, y_top-0.44, f"В{p.variant}-П{p.replication}",
                                ha="center", va="center", fontsize=8,
                                fontfamily="Times New Roman", color="#1a4b8c", fontweight="bold")
                        bot_r = matplotlib.patches.Rectangle(
                            (x+0.03, y_top-1.55), 0.94, 0.78,
                            facecolor="white", edgecolor="#333", linewidth=1.2)
                        ax.add_patch(bot_r)
                    else:
                        label = {HP_ROLE_GUARD_EDGE:"K", HP_ROLE_GUARD_REP:"P",
                                 HP_ROLE_DEAD:"-", HP_ROLE_POLLINIZER:"+",
                                 HP_ROLE_EXTRA:"x"}.get(p.role, "")
                        rect = matplotlib.patches.Rectangle(
                            (x+0.12, y_top-1.2), 0.76, 1.0,
                            facecolor="#eeeeee", edgecolor="#bbb", linewidth=0.6)
                        ax.add_patch(rect)
                        if label:
                            ax.text(x+0.5, y_top-0.7, label, ha="center", va="center",
                                    fontsize=8, color="#999", fontfamily="Times New Roman")
            ax.set_xlim(-1.6, max_len+0.5); ax.set_ylim(0, n_seg*ROW_H+1.3)
            ax.axis("off")
            ax.set_title(
                "БЛАНК ОБЛІКУ\n"
                "Показник: _______________________   Одиниця: _________\n"
                f"Дата: _______________     Виконав: _______________________     "
                f"Сторінка {page_idx['i']+1}/{len(pages)}",
                fontsize=10, fontfamily="Times New Roman", loc="left")
            fig.subplots_adjust(top=0.82, bottom=0.03, left=0.1, right=0.98)
            state["fig"] = fig
            lbl.configure(text=f"Сторінка {page_idx['i']+1} / {len(pages)}")
            embed_figure(fig, holder)

        def _step(d):
            page_idx["i"] = max(0, min(len(pages)-1, page_idx["i"]+d))
            _render()

        tk.Button(tb, text="◀ Попередня", font=("Times New Roman",10),
                  command=lambda: _step(-1)).pack(side=tk.LEFT, padx=4)
        tk.Button(tb, text="Наступна ▶", font=("Times New Roman",10),
                  command=lambda: _step(1)).pack(side=tk.LEFT, padx=4)
        def _save():
            if state["fig"] is None: return
            path = filedialog.asksaveasfilename(defaultextension=".png",
                        filetypes=[("PNG зображення","*.png")],
                        initialfile=f"blank_form_page_{page_idx['i']+1}.png",
                        title="Зберегти сторінку")
            if not path: return
            try:
                state["fig"].savefig(path, dpi=150, bbox_inches="tight")
                messagebox.showinfo("Збережено", f"Збережено:\n{path}")
            except Exception as ex:
                messagebox.showerror("Помилка", str(ex))
        tk.Button(tb, text="💾 Зберегти PNG", font=("Times New Roman",10),
                  command=_save).pack(side=tk.LEFT, padx=12)
        _render()

    # ── Зведення журналу в таблицю для аналізу ─────────────
    def _open_aggregate_dialog(self):
        if not self.plants:
            messagebox.showwarning("", "Спочатку відкрийте схему або журнал."); return
        if not self.records:
            messagebox.showwarning("", "У журналі ще немає жодного обліку — "
                                       "спершу створіть облік і внесіть значення."); return
        self._save_current_values_from_widgets()

        dlg = tk.Toplevel(self.win); dlg.title("Звести дані для аналізу")
        dlg.resizable(False, False); set_icon(dlg); dlg.grab_set()
        frm = tk.Frame(dlg, padx=16, pady=14); frm.pack()
        tk.Label(frm,
                 text="Оберіть один або кілька обліків (показників), які потрібно звести в\n"
                      "підсумкову таблицю. Якщо обрати кілька (наприклад, «Урожайність 2022»,\n"
                      "«2023», «2024») — вони стануть окремими стовпцями однієї таблиці,\n"
                      "придатної для аналізу динаміки в часі (Змішаний RM).",
                 font=("Times New Roman",10), fg="#555", justify="left"
                 ).pack(anchor="w", pady=(0,10))
        check_vars = {}
        for nm in self.records.keys():
            v = tk.BooleanVar(value=True)
            tk.Checkbutton(frm, text=nm, variable=v, font=("Times New Roman",11)
                           ).pack(anchor="w")
            check_vars[nm] = v

        def _go():
            chosen = [nm for nm, v in check_vars.items() if v.get()]
            if not chosen:
                messagebox.showwarning("", "Оберіть хоча б один облік.", parent=dlg); return
            dlg.destroy()
            self._show_aggregate_result(chosen)
        bf = tk.Frame(frm); bf.pack(pady=(12,0))
        tk.Button(bf, text="Звести →", bg="#1a6b1a", fg="white",
                  font=("Times New Roman",11), command=_go).pack(side=tk.LEFT, padx=4)
        tk.Button(bf, text="Скасувати", font=("Times New Roman",11),
                  command=dlg.destroy).pack(side=tk.LEFT)
        center_win(dlg)

    def _aggregate_for_analysis(self, record_names):
        """Зводить журнал у таблицю: одна строка на КОЖНУ повторність (не на
        кожну облікову рослину!) — значення всіх облікових рослин В МЕЖАХ
        однієї повторності усереднюються в одне число. Це методично
        правильно: облікові рослини всередині повторності — субпроби
        однієї дослідної одиниці, а не незалежні повторення, тож пряме
        підставлення кожної окремо в ANOVA спричинило б псевдоповторність.

        Підтримує і однофакторні схеми (self.variant), і багатофакторні
        (p.factors — окремий рівень для кожного фактора одночасно).
        Делегує спільній функції модуля (той самий код, що й у
        «Відкрити показник» з інших аналізів)."""
        _factor_cols, rows = aggregate_journal_data(
            self.plants, self.factor_defs, self.variant_names, self.records, record_names)
        return rows

    def _pivot_wide(self, record_name, rows, factor_cols):
        """Перетворює довгий список (комбінація×повторність) на ШИРОКИЙ
        формат — саме такий, як в таблиці вводу даних ANOVA: рядок на
        унікальну комбінацію рівнів факторів (фактор А, фактор Б…),
        стовпці Повт.1, Повт.2… Кожна облікова рослина вже закодована
        в схемі своїм варіантом і повторністю, тож ця функція просто
        розкладає вже усереднені (в межах повторності) значення по
        правильних місцях таблиці."""
        by_combo = {}
        for r in rows:
            combo_key = tuple(r[fc] for fc in factor_cols)
            by_combo.setdefault(combo_key, {})[r["replication"]] = r[record_name]
        combos_sorted = sorted(by_combo.keys())
        all_reps = sorted({rep for v in by_combo.values() for rep in v.keys()})
        headers = factor_cols + [f"Повт.{i+1}" for i in range(len(all_reps))]
        tbl_rows = []
        for combo in combos_sorted:
            row = list(combo)
            rep_vals = by_combo[combo]
            row += ["" if rep_vals.get(rep) is None else rep_vals.get(rep) for rep in all_reps]
            tbl_rows.append(row)
        return headers, tbl_rows

    def _show_aggregate_result(self, record_names):
        rows = self._aggregate_for_analysis(record_names)
        if not rows:
            messagebox.showwarning("", "Немає жодної сформованої повторності "
                                       "(облікові рослини не визначені в схемі)."); return
        is_multi = bool(self.factor_defs) and any(p.factors for p in self.plants)
        factor_cols = [fdef["name"] for fdef in self.factor_defs] if is_multi else ["Варіант"]

        win = tk.Toplevel(self.win); win.title("Зведена таблиця для аналізу")
        win.geometry("980x600"); set_icon(win)
        tb = tk.Frame(win, padx=6, pady=5); tb.pack(fill=tk.X)

        if len(record_names) == 1:
            # ОДИН показник — одразу широкий формат: фактор А, фактор Б…,
            # Повт.1, Повт.2… Кожна облікова рослина вже має в схемі свій
            # варіант і повторність, тому дані одразу стають у потрібні
            # клітинки без ручного перекладання.
            headers, tbl_rows = self._pivot_wide(record_names[0], rows, factor_cols)
            tk.Button(tb, text="📋 Копіювати таблицю", font=("Times New Roman",11),
                      command=lambda: self._copy_table(win, headers, tbl_rows)
                      ).pack(side=tk.LEFT, padx=4)
            if len(factor_cols) <= 4:
                tk.Button(tb, text="➡ Відкрити в головній ANOVA-таблиці", bg="#1a6b1a", fg="white",
                          font=("Times New Roman",11),
                          command=lambda: self._open_in_main_anova(record_names[0], rows, factor_cols)
                          ).pack(side=tk.LEFT, padx=4)
            note = ("Кожен рядок — унікальна комбінація " +
                    ("рівнів факторів" if is_multi else "варіанту") +
                    "; кожен стовпець «Повт.N» — середнє значення показника САМЕ в цій "
                    "повторності (усереднено з облікових рослин у її межах — уникає "
                    "псевдоповторності). Формат готовий для прямої вставки в ANOVA-таблицю.")
        else:
            # Кілька показників одночасно (напр. кілька років) — широкий
            # формат тут природний для ІНШОЇ мети (Змішаний RM: варіант +
            # повторність в рядку, показники — стовпцями), тож таблиця
            # лишається довгою (рядок = комбінація × повторність).
            headers = factor_cols + ["Повторність","К-сть субпроб"] + record_names
            tbl_rows = [[r[fc] for fc in factor_cols] + [r["replication"], r["n_subsamples"]] +
                        [("" if r[rn] is None else r[rn]) for rn in record_names] for r in rows]
            tk.Button(tb, text="📋 Копіювати таблицю", font=("Times New Roman",11),
                      command=lambda: self._copy_table(win, headers, tbl_rows)
                      ).pack(side=tk.LEFT, padx=4)
            if not is_multi:
                tk.Button(tb, text="➡ Відкрити в «Змішаний RM»", bg="#1a6b1a", fg="white",
                          font=("Times New Roman",11),
                          command=lambda: self._open_in_mixed_rm(record_names, rows)
                          ).pack(side=tk.LEFT, padx=4)
            note = ("Кілька показників одночасно (напр. кілька років) показуються в "
                    "довгому форматі — рядок на комбінацію × повторність, оскільки "
                    "широкий формат по повторностях природний лише для ОДНОГО показника "
                    "за раз.")

        tk.Label(win, text=note, font=("Times New Roman",10), fg="#555", justify="left",
                 wraplength=940, anchor="w").pack(fill=tk.X, padx=10, pady=(4,4))
        frm, _ = make_tv(win, headers, tbl_rows)
        frm.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0,10))

    def _copy_table(self, win, headers, tbl_rows):
        lines = ["\t".join(headers)]
        for row in tbl_rows:
            lines.append("\t".join(str(v) for v in row))
        win.clipboard_clear(); win.clipboard_append("\n".join(lines))
        messagebox.showinfo("Скопійовано",
            "Таблицю скопійовано у буфер обміну.\n"
            "Вставте (Ctrl+V) у таблицю потрібного аналізу — курсор поставте "
            "в перший стовпчик, перший рядок.")

    def _open_in_mixed_rm(self, record_names, rows):
        """Відкриває Змішаний RM і одразу заповнює його таблицю зведеними
        даними — Варіант/Повторність збігаються за форматом один-в-один."""
        w = MixedRepeatedWindow(self._parent, self.gs)
        while len(w.time_vars) < len(record_names): w._add_col()
        while len(w.time_vars) > len(record_names): w._del_col()
        for i, rn in enumerate(record_names):
            w.time_vars[i].set(rn)
        while len(w.entries) < len(rows): w._add_row()
        while len(w.entries) > len(rows): w._del_row()
        for i, r in enumerate(rows):
            w.entries[i][0].delete(0, tk.END); w.entries[i][0].insert(0, str(r["Варіант"]))
            w.entries[i][1].delete(0, tk.END); w.entries[i][1].insert(0, str(r["replication"]))
            for j, rn in enumerate(record_names):
                val = r[rn]
                w.entries[i][2+j].delete(0, tk.END)
                if val is not None: w.entries[i][2+j].insert(0, str(val))
        messagebox.showinfo("Дані перенесено",
            f"Дані з {len(record_names)} обліків перенесено у «Змішаний RM»\n"
            f"({len(rows)} повторностей). Перевірте таблицю і натисніть «▶ Аналіз».")

    def _open_in_main_anova(self, record_name, rows, factor_cols):
        """Відкриває головну багатофакторну ANOVA-таблицю програми (до 4
        факторів) і одразу заповнює її: один рядок на кожну унікальну
        комбінацію рівнів факторів, значення повторностей — у стовпці
        Повт.1, Повт.2… (номер визначається порядком номерів повторності)."""
        app = getattr(self._parent, "_sad_app", None)
        if app is None:
            messagebox.showerror("Не вдалося",
                "Не знайдено головного вікна програми — відкрийте таблицю "
                "вручну і скористайтесь «📋 Копіювати»."); return
        fc = len(factor_cols)
        if fc > 4:
            messagebox.showwarning("", "Головна таблиця підтримує максимум "
                                       "4 фактори одночасно."); return

        by_combo = {}
        for r in rows:
            combo_key = tuple(r[fname] for fname in factor_cols)
            by_combo.setdefault(combo_key, {})[r["replication"]] = r[record_name]
        combos_sorted = sorted(by_combo.keys())
        all_reps = sorted({rep for v in by_combo.values() for rep in v.keys()})
        n_rep_cols = max(len(all_reps), 1)

        app.open_table(fc)
        for i, fname in enumerate(factor_cols):
            key = app.factor_keys[i]
            app._set_ftitle(key, fname)
            app.header_labels[i].configure(text=app.ftitle(key))

        while len(app.entries) < len(combos_sorted): app.add_row()
        while (app.cols - fc) < n_rep_cols: app.add_column()

        for i, combo in enumerate(combos_sorted):
            for k, lvl_name in enumerate(combo):
                app.entries[i][k].delete(0, tk.END)
                app.entries[i][k].insert(0, str(lvl_name))
            rep_vals = by_combo[combo]
            for ci, rep_num in enumerate(all_reps):
                col = fc + ci
                if col >= app.cols: continue
                v = rep_vals.get(rep_num)
                app.entries[i][col].delete(0, tk.END)
                if v is not None: app.entries[i][col].insert(0, str(v))

        messagebox.showinfo("Дані перенесено",
            f"Дані показника «{record_name}» перенесено в головну {fc}-факторну "
            f"ANOVA-таблицю ({len(combos_sorted)} комбінацій × {n_rep_cols} повторень). "
            "Перевірте таблицю і натисніть «▶ Аналіз».")


# ═══════════════════════════════════════════════════════════════
# TRIAL DESIGN GENERATOR
# ═══════════════════════════════════════════════════════════════
class TrialDesignWindow:
    """Генератор плану польового досліду — універсальний для всіх культур."""

    # ── Типи культур з налаштуваннями ────────────────────────
    CULTURES = {
        "Зернові / польові культури": {
            "plot_w": 3.0, "plot_l": 10.0, "unit": "ділянка",
            "garden": False,
            "indicators": ["Висота рослин, см", "Маса 1000 зерен, г",
                           "Врожайність, т/га", "Вміст білку, %"],
        },
        "Садівництво (дерева)": {
            "plot_w": 4.0, "plot_l": 5.0, "unit": "дерево",
            "garden": True,
            "row_sp": 4.0, "plant_sp": 5.0,
            "plants_plot": 5, "guard_ends": 1, "guard_rows": 1,
            "indicators": ["Висота дерева, м", "Діаметр крони, м",
                           "Врожайність з дерева, кг", "Маса плоду, г",
                           "Вміст ЦРР, °Brix", "% зав'язування квіток"],
        },
        "Ягідники": {
            "plot_w": 2.0, "plot_l": 0.5, "unit": "кущ",
            "garden": True,
            "row_sp": 2.0, "plant_sp": 0.5,
            "plants_plot": 8, "guard_ends": 2, "guard_rows": 1,
            "indicators": ["Висота куща, см", "Кількість пагонів",
                           "Врожайність з куща, кг", "Маса ягоди, г",
                           "Вміст ЦРР, °Brix"],
        },
        "Овочівництво (відкритий ґрунт)": {
            "plot_w": 2.0, "plot_l": 5.0, "unit": "ділянка",
            "garden": False,
            "indicators": ["Висота рослин, см", "Маса плоду, г",
                           "Врожайність, т/га", "Товарність, %",
                           "Вихід стандарту, %"],
        },
        "Захищений ґрунт (теплиця)": {
            "plot_w": 1.0, "plot_l": 4.0, "unit": "грядка",
            "garden": False,
            "indicators": ["Висота рослин, см", "Кількість плодів/рослину",
                           "Врожайність, кг/м²", "Маса плоду, г"],
        },
        "Виноградарство": {
            "plot_w": 3.0, "plot_l": 1.5, "unit": "кущ",
            "garden": True,
            "row_sp": 3.0, "plant_sp": 1.5,
            "plants_plot": 6, "guard_ends": 2, "guard_rows": 1,
            "indicators": ["Маса грона, г", "Кількість грон/кущ",
                           "Врожайність з куща, кг", "Вміст ЦРР, °Brix",
                           "Кислотність, г/л"],
        },
    }

    DESIGNS = [
        ("crd",   "CRD — Повністю рандомізований",
         "Всі ділянки рівноцінні. Варіанти розміщуються випадково."),
        ("rcbd",  "RCBD — Рандомізовані повні блоки (рекомендується)",
         "Поле ділиться на повторності. Кожна повторність = всі варіанти."),
        ("latin", "Латинський квадрат",
         "Контролює 2 джерела мінливості. k варіантів = k рядів = k стовпців."),
        ("split", "Split-plot — Розщеплені ділянки",
         "Два фактори різного масштабу. WP = великі ділянки, SP = підділянки."),
    ]

    HELP_TEXT = """
ГЕНЕРАТОР ПЛАНУ ПОЛЬОВОГО ДОСЛІДУ
══════════════════════════════════════════════════

ЩО РОБИТЬ?
  Автоматично рандомізує розміщення варіантів
  і формує документи для польової роботи:
  • Польова схема (кольорова карта ділянок)
  • Список рандомізації (порядок закладки)
  • Польовий журнал (таблиця для вимірювань)

══════════════════════════════════════════════════
КРОК 1. ТИП КУЛЬТУРИ
══════════════════════════════════════════════════
  Оберіть тип культури — програма підлаштує:
  • Типові розміри ділянок
  • Стандартні показники для журналу
  • Термінологію (ділянка/дерево/кущ)

══════════════════════════════════════════════════
КРОК 2. ВАРІАНТИ ДОСЛІДУ
══════════════════════════════════════════════════
  Введіть назви варіантів — по одному на рядок.
  Приклади:
    Контроль (без добрив)
    N60P60K60
    N90P60K60
    N120P60K60

  Для Split-plot — введіть також sub-plot варіанти.

══════════════════════════════════════════════════
КРОК 3. ДИЗАЙН ДОСЛІДУ
══════════════════════════════════════════════════

  CRD — Повністю рандомізований:
    Для однорідних умов. Всі ділянки рівноцінні.
    Варіанти розміщуються абсолютно випадково.
    Простий але потребує однорідного фону.

  RCBD — Рандомізовані повні блоки:
    РЕКОМЕНДУЄТЬСЯ для більшості дослідів.
    Поле ділиться на ПОВТОРНОСТІ (блоки).
    Кожна повторність містить всі варіанти.
    Блоки розміщують перпендикулярно до основного
    градієнта мінливості (схил, зрошення, ряди).
    У садівництві: повторність = кілька дерев/ряд.

  Латинський квадрат:
    Контролює ДВА незалежних джерела мінливості.
    k варіантів → k рядів × k стовпців ділянок.
    Кожен варіант — рівно 1 раз у кожному ряду
    і рівно 1 раз у кожному стовпці.
    Рекомендується при k=4-6 варіантів.
    Після досліду → 3-факторна ANOVA у S.A.D.
    (Фактор A = Варіант, B = Рядок, C = Стовпець)

  Split-plot — Розщеплені ділянки:
    Для двох факторів різного масштабу.
    WP (whole-plot): головний фактор — великі ділянки.
    SP (sub-plot): другорядний — всередині WP.
    Приклади:
      Зернові:    WP = обробка ґрунту, SP = сорт
      Садівництво: WP = підщепа, SP = сорт
      Овочі:      WP = спосіб вирощування, SP = сорт

══════════════════════════════════════════════════
КРОК 4. SEED РАНДОМІЗАЦІЇ
══════════════════════════════════════════════════
  Seed — технічний номер жеребкування.
  При однаковому seed → однакова схема.
  ЗБЕРІГАЙТЕ seed у документації досліду!
  Seed ≠ Рік: seed може бути будь-яким числом.

══════════════════════════════════════════════════
КРОК 5. ПОЛЬОВИЙ ЖУРНАЛ
══════════════════════════════════════════════════
  Налаштуйте назви показників (через ";").
  Програма підставляє типові показники для
  обраного типу культури.
  Натисніть "▶ Оновити" після зміни назв.
  Збережіть журнал у Excel для польової роботи.

══════════════════════════════════════════════════
САДІВНИЦТВО — ОСОБЛИВОСТІ
══════════════════════════════════════════════════
  Ділянка = кілька дерев одного варіанту в ряду.
  Повторність = окремий ряд або частина ряду.

  Приклад для яблуні (схема 4×5 м):
    4 сорти × 3 повторності = 12 ділянок
    По 3-5 дерев на ділянку
    Розмір ділянки: 4 м (ширина) × 15 м (5 дерев)

  У RCBD: кожна повторність = 1 ряд дерев,
  де всі 4 сорти розміщені випадково в ряду.

  Показники для кожного дерева усереднюють
  і записують одне середнє значення на ділянку.
"""

    def __init__(self, parent):
        self.win = tk.Toplevel(parent)
        self.win.title("Генератор плану польового досліду")
        self.win.geometry("1160x760")
        self.win.resizable(True, True)
        set_icon(self.win)
        self._plan_data = None
        self._build()

    # ═══════════════════════════════════════════════════════
    # _build — головний інтерфейс
    # ═══════════════════════════════════════════════════════
    def _build(self):
        rf = ("Times New Roman", 11)

        # ── Toolbar ────────────────────────────────────────
        top = tk.Frame(self.win, padx=8, pady=5); top.pack(fill=tk.X)
        tk.Button(top, text="▶ Згенерувати план", bg="#c62828", fg="white",
                  font=("Times New Roman", 13),
                  command=self._generate).pack(side=tk.LEFT, padx=4)
        tk.Button(top, text="📚 Довідка", bg="#1a4b8c", fg="white",
                  font=rf, command=self._show_help).pack(side=tk.LEFT, padx=4)

        # ── Основна область ────────────────────────────────
        main = tk.Frame(self.win); main.pack(fill=tk.BOTH, expand=True, padx=8, pady=4)

        # ── ЛІВА ПАНЕЛЬ (прокручувана) ─────────────────────
        left_outer = tk.Frame(main, width=360)
        left_outer.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 8))
        left_outer.pack_propagate(False)
        lc = tk.Canvas(left_outer, highlightthickness=0)
        lc.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        lsb = ttk.Scrollbar(left_outer, orient="vertical", command=lc.yview)
        lsb.pack(side=tk.RIGHT, fill=tk.Y)
        lc.configure(yscrollcommand=lsb.set)
        lf = tk.Frame(lc); lc.create_window((0, 0), window=lf, anchor="nw")
        lf.bind("<Configure>", lambda e: lc.configure(scrollregion=lc.bbox("all")))
        left_outer.bind("<MouseWheel>",
                        lambda e: lc.yview_scroll(int(-1*(e.delta/120)), "units"))

        # ─── Тип культури ──────────────────────────────────
        cf = tk.LabelFrame(lf, text="1. Тип культури",
                           font=("Times New Roman", 11, "bold"), padx=8, pady=4)
        cf.pack(fill=tk.X, pady=(0, 4))
        self.culture_var = tk.StringVar(value=list(self.CULTURES.keys())[0])
        self.culture_cb = ttk.Combobox(cf, textvariable=self.culture_var,
                                       values=list(self.CULTURES.keys()),
                                       state="readonly", width=36, font=rf)
        self.culture_cb.pack(fill=tk.X, pady=2)
        self.culture_cb.bind("<<ComboboxSelected>>", self._on_culture)
        self._culture_hint = tk.Label(cf, text="", font=("Times New Roman", 9),
                                      fg="#555", justify="left")
        self._culture_hint.pack(anchor="w")

        # ─── Варіанти ──────────────────────────────────────
        vf = tk.LabelFrame(lf, text="2. Варіанти досліду (один на рядок)",
                           font=("Times New Roman", 11, "bold"), padx=8, pady=4)
        vf.pack(fill=tk.X, pady=(0, 4))
        self.var_text = tk.Text(vf, width=38, height=6, font=rf, wrap="word")
        self.var_text.pack(fill=tk.X, pady=2)
        self.var_text.insert("1.0", "Контроль\nВаріант 1\nВаріант 2\nВаріант 3")

        # ─── Дизайн ────────────────────────────────────────
        df = tk.LabelFrame(lf, text="3. Дизайн досліду",
                           font=("Times New Roman", 11, "bold"), padx=8, pady=4)
        df.pack(fill=tk.X, pady=(0, 4))
        self.design_var = tk.StringVar(value="rcbd")
        for val, label, desc in self.DESIGNS:
            fr = tk.Frame(df); fr.pack(fill=tk.X, pady=1)
            tk.Radiobutton(fr, text=label, variable=self.design_var, value=val,
                           font=("Times New Roman", 11),
                           command=self._on_design).pack(side=tk.LEFT)
        self._design_hint = tk.Label(df, text="", font=("Times New Roman", 9),
                                     fg="#1a4b8c", bg="#eef4ff",
                                     justify="left", wraplength=320, padx=4, pady=3)
        self._design_hint.pack(fill=tk.X, pady=4)

        # Split-plot додатковий фактор
        self.sp_frame = tk.LabelFrame(lf, text="Sub-plot варіанти",
                                      font=("Times New Roman", 11, "bold"),
                                      padx=8, pady=4)
        self.sp_text = tk.Text(self.sp_frame, width=38, height=3, font=rf)
        self.sp_text.pack(fill=tk.X)
        self.sp_text.insert("1.0", "Сорт А\nСорт Б\nСорт В")
        # (показується лише для split-plot)

        # ─── Параметри ─────────────────────────────────────
        pf = tk.LabelFrame(lf, text="4. Параметри",
                           font=("Times New Roman", 11, "bold"), padx=8, pady=4)
        pf.pack(fill=tk.X, pady=(0, 4))
        self._pv = {}
        for ri, (lbl, key, default, hint) in enumerate([
            ("Повторностей:", "reps", "3",
             "Кількість повторностей (рядів)"),
            ("Seed рандомізації:", "seed", "2024",
             "Число для відтворення жеребкування. ≠ рік!"),
        ]):
            tk.Label(pf, text=lbl, font=rf).grid(row=ri, column=0, sticky="w", pady=2)
            v = tk.StringVar(value=default); self._pv[key] = v
            tk.Entry(pf, textvariable=v, width=9, font=rf
                     ).grid(row=ri, column=1, sticky="w", padx=6)
            tk.Label(pf, text=hint, font=("Times New Roman", 9), fg="#666"
                     ).grid(row=ri, column=2, sticky="w")

        # ─── Польові параметри (для зернових/овочів) ────────
        self._field_frame = tk.LabelFrame(lf, text="Розміри ділянки",
                           font=("Times New Roman", 11, "bold"), padx=8, pady=4)
        self._field_frame.pack(fill=tk.X, pady=(0, 4))
        for ri, (lbl, key, default, hint) in enumerate([
            ("Ширина, м:", "pw", "5", "Ширина ділянки"),
            ("Довжина, м:", "pl", "10", "Довжина ділянки"),
        ]):
            tk.Label(self._field_frame, text=lbl, font=rf
                     ).grid(row=ri, column=0, sticky="w", pady=2)
            v = tk.StringVar(value=default); self._pv[key] = v
            tk.Entry(self._field_frame, textvariable=v, width=9, font=rf
                     ).grid(row=ri, column=1, sticky="w", padx=6)
            tk.Label(self._field_frame, text=hint,
                     font=("Times New Roman", 9), fg="#666"
                     ).grid(row=ri, column=2, sticky="w")

        # ─── Садівничі параметри ─────────────────────────────
        self._garden_frame = tk.LabelFrame(lf, text="Параметри садіння",
                            font=("Times New Roman", 11, "bold"), padx=8, pady=4)
        self._gv = {}
        garden_params = [
            ("Між рядами (A), м:",              "row_sp",     "4.0",
             "Відстань між рядами"),
            ("В ряду (B), м:",                  "plant_sp",   "5.0",
             "Відстань між рослинами в ряду"),
            ("Облікових рослин на ділянку:",   "plants_plot","5",
             "Без захисних — лише облікові"),
            ("Захисних рослин (поч. і кін.):", "guard_ends", "1",
             "Кількість з кожного боку ряду"),
            ("Захисних рядів між варіантами:", "guard_rows", "1",
             "Рядів-буферів між повторностями"),
        ]
        for ri, (lbl, key, default, hint) in enumerate(garden_params):
            tk.Label(self._garden_frame, text=lbl, font=rf
                     ).grid(row=ri, column=0, sticky="w", pady=2)
            v = tk.StringVar(value=default); self._gv[key] = v
            tk.Entry(self._garden_frame, textvariable=v, width=9, font=rf
                     ).grid(row=ri, column=1, sticky="w", padx=6)
            tk.Label(self._garden_frame, text=hint,
                     font=("Times New Roman", 9), fg="#666"
                     ).grid(row=ri, column=2, sticky="w")

        # ─── Паспорт ───────────────────────────────────────
        nf = tk.LabelFrame(lf, text="5. Паспорт досліду",
                           font=("Times New Roman", 11, "bold"), padx=8, pady=4)
        nf.pack(fill=tk.X, pady=(0, 4))
        self._nv = {}
        for ri, (lbl, key) in enumerate([
            ("Назва:", "name"), ("Рік:", "year"),
            ("Місце:", "loc"),  ("Відповідальний:", "resp"),
        ]):
            tk.Label(nf, text=lbl, font=rf).grid(row=ri, column=0, sticky="w", pady=2)
            v = tk.StringVar(value=(str(datetime.now().year) if key=="year" else ""))
            self._nv[key] = v
            tk.Entry(nf, textvariable=v, width=28, font=rf
                     ).grid(row=ri, column=1, sticky="w", padx=6, pady=2)

        # ── ПРАВА ПАНЕЛЬ — вкладки результатів ─────────────
        right = tk.Frame(main); right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.nb = ttk.Notebook(right); self.nb.pack(fill=tk.BOTH, expand=True)

        # Вкладка 1: Схема
        t1 = tk.Frame(self.nb); self.nb.add(t1, text="🗺 Польова схема")
        t1_tb = tk.Frame(t1); t1_tb.pack(fill=tk.X, padx=4, pady=3)
        tk.Label(t1_tb,
                 text="Клітинки = ділянки/рослини. Кольори = варіанти.",
                 font=("Times New Roman", 9), fg="#666").pack(side=tk.LEFT)
        tk.Button(t1_tb, text="💾 Зберегти PNG", font=("Times New Roman",10),
                  command=self._save_png).pack(side=tk.RIGHT, padx=4)
        self._scheme_cv = tk.Canvas(t1, bg="white")
        s_vsb = ttk.Scrollbar(t1, orient="vertical",
                               command=self._scheme_cv.yview)
        s_hsb = ttk.Scrollbar(t1, orient="horizontal",
                               command=self._scheme_cv.xview)
        s_vsb.pack(side=tk.RIGHT, fill=tk.Y)
        s_hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self._scheme_cv.pack(fill=tk.BOTH, expand=True)
        self._scheme_cv.configure(yscrollcommand=s_vsb.set,
                                   xscrollcommand=s_hsb.set)

        # Вкладка 2: Рандомізація
        t2 = tk.Frame(self.nb); self.nb.add(t2, text="📋 Рандомізація")
        tb2 = tk.Frame(t2); tb2.pack(fill=tk.X, padx=4, pady=3)
        tk.Label(tb2, text="Порядок закладки ділянок:",
                 font=rf).pack(side=tk.LEFT)
        tk.Button(tb2, text="💾 Зберегти TXT", font=("Times New Roman",10),
                  command=self._save_rand_txt).pack(side=tk.RIGHT, padx=4)
        r_vsb = ttk.Scrollbar(t2, orient="vertical")
        r_vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.rand_txt = tk.Text(t2, font=("Courier New", 10),
                                yscrollcommand=r_vsb.set,
                                state="disabled", wrap="none")
        self.rand_txt.pack(fill=tk.BOTH, expand=True)
        r_vsb.config(command=self.rand_txt.yview)

        # Вкладка 3: Журнал
        t3 = tk.Frame(self.nb); self.nb.add(t3, text="📓 Польовий журнал")
        tb3 = tk.Frame(t3); tb3.pack(fill=tk.X, padx=4, pady=3)
        tk.Button(tb3, text="💾 Зберегти Excel", font=("Times New Roman",10),
                  command=self._save_excel).pack(side=tk.RIGHT, padx=4)
        tk.Label(tb3, text="Показники:", font=rf).pack(side=tk.LEFT, padx=(0,4))
        self.ind_var = tk.StringVar()
        tk.Entry(tb3, textvariable=self.ind_var, width=55,
                 font=("Times New Roman", 10)).pack(side=tk.LEFT, padx=2)
        tk.Button(tb3, text="▶ Оновити", bg="#c62828", fg="white",
                  font=("Times New Roman", 10),
                  command=self._refresh_journal).pack(side=tk.LEFT, padx=4)
        tk.Label(tb3, text="(через крапку з комою)",
                 font=("Times New Roman", 9), fg="#888").pack(side=tk.LEFT)

        j_frame = tk.Frame(t3); j_frame.pack(fill=tk.BOTH, expand=True)
        j_vsb = ttk.Scrollbar(j_frame, orient="vertical")
        j_hsb = ttk.Scrollbar(j_frame, orient="horizontal")
        j_vsb.pack(side=tk.RIGHT, fill=tk.Y)
        j_hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self.journal_tv = ttk.Treeview(j_frame, yscrollcommand=j_vsb.set,
                                       xscrollcommand=j_hsb.set)
        self.journal_tv.pack(fill=tk.BOTH, expand=True)
        j_vsb.config(command=self.journal_tv.yview)
        j_hsb.config(command=self.journal_tv.xview)

        # Ініціалізуємо підказки
        self._on_culture()
        self._on_design()

    # ═══════════════════════════════════════════════════════
    # Обробники змін
    # ═══════════════════════════════════════════════════════
    def _on_culture(self, *_):
        key = self.culture_var.get()
        cfg = self.CULTURES.get(key, {})
        unit = cfg.get("unit", "ділянка")
        is_garden = cfg.get("garden", False)

        if is_garden:
            row_sp    = cfg.get("row_sp",    4.0)
            plant_sp  = cfg.get("plant_sp",  5.0)
            n_plot    = cfg.get("plants_plot",5)
            g_ends    = cfg.get("guard_ends", 1)
            g_rows    = cfg.get("guard_rows", 1)
            self._culture_hint.configure(
                text=f"Одиниця: {unit}  |  Схема: {row_sp}×{plant_sp} м")
            if hasattr(self, '_gv'):
                # Підказки заповнюємо лише якщо поля порожні
                if not self._gv["row_sp"].get() or self._gv["row_sp"].get() == "0":
                    self._gv["row_sp"].set(str(row_sp))
                if not self._gv["plant_sp"].get() or self._gv["plant_sp"].get() == "0":
                    self._gv["plant_sp"].set(str(plant_sp))
                if not self._gv["plants_plot"].get():
                    self._gv["plants_plot"].set(str(n_plot))
                if not self._gv["guard_ends"].get():
                    self._gv["guard_ends"].set(str(g_ends))
                if not self._gv["guard_rows"].get():
                    self._gv["guard_rows"].set(str(g_rows))
            # Показати садівничий фрейм, сховати польовий
            if hasattr(self, '_garden_frame'):
                self._field_frame.pack_forget()
                self._garden_frame.pack(fill=tk.X, pady=(0,4),
                                        after=self._field_frame)
        else:
            pw = cfg.get("plot_w", 5); pl = cfg.get("plot_l", 10)
            self._culture_hint.configure(
                text=f"Одиниця: {unit}  |  Типові розміри: {pw}×{pl} м")
            if hasattr(self, '_pv'):
                self._pv["pw"].set(str(pw))
                self._pv["pl"].set(str(pl))
            # Показати польовий фрейм, сховати садівничий
            if hasattr(self, '_garden_frame'):
                self._garden_frame.pack_forget()
                self._field_frame.pack(fill=tk.X, pady=(0,4))

        indicators = cfg.get("indicators", [])
        if hasattr(self, 'ind_var'):
            self.ind_var.set("; ".join(indicators))

    def _on_design(self, *_):
        val = self.design_var.get()
        hints = {
            "crd":   "Однорідний фон. Варіанти розміщуються абсолютно випадково по всьому полю.",
            "rcbd":  "Рекомендується. Поле ділиться на повторності. Кожна повторність = всі варіанти.",
            "latin": "k варіантів = k рядів = k стовпців. Аналіз — 3-факторна ANOVA у S.A.D.",
            "split": "Введіть WP (whole-plot) варіанти вище і SP варіанти нижче.",
        }
        if hasattr(self, '_design_hint'):
            self._design_hint.configure(text=hints.get(val, ""))
        if hasattr(self, 'sp_frame'):
            if val == "split":
                self.sp_frame.pack(fill=tk.X, pady=(0, 4),
                                   after=self.sp_frame.master.winfo_children()[-1]
                                   if self.sp_frame.master.winfo_children() else None)
            else:
                self.sp_frame.pack_forget()

    # ═══════════════════════════════════════════════════════
    # Довідка
    # ═══════════════════════════════════════════════════════
    def _show_help(self):
        win = tk.Toplevel(self.win)
        win.title("Довідка — Генератор плану досліду")
        win.geometry("700x660"); set_icon(win)
        frm = tk.Frame(win); frm.pack(fill=tk.BOTH, expand=True, padx=8, pady=6)
        vsb = ttk.Scrollbar(frm, orient="vertical"); vsb.pack(side=tk.RIGHT, fill=tk.Y)
        txt = tk.Text(frm, wrap="word", font=("Times New Roman", 11),
                      yscrollcommand=vsb.set, relief=tk.FLAT,
                      bg="#fafafa", padx=10, pady=8, cursor="arrow")
        txt.pack(fill=tk.BOTH, expand=True); vsb.config(command=txt.yview)
        txt.insert("1.0", self.HELP_TEXT.strip()); txt.configure(state="disabled")
        txt.bind("<MouseWheel>",
                 lambda e: txt.yview_scroll(int(-1*(e.delta/120)), "units"))
        tk.Button(win, text="Закрити", command=win.destroy,
                  font=("Times New Roman", 11)).pack(pady=6)

    # ═══════════════════════════════════════════════════════
    # Генерація плану
    # ═══════════════════════════════════════════════════════
    def _generate(self):
        import random

        variants = [v.strip() for v in
                    self.var_text.get("1.0", "end").splitlines() if v.strip()]
        if len(variants) < 2:
            messagebox.showwarning("", "Введіть щонайменше 2 варіанти."); return

        try:
            reps = int(self._pv["reps"].get())
            seed = int(self._pv["seed"].get())
        except ValueError:
            messagebox.showwarning("", "Перевірте числові поля."); return

        cfg = self.CULTURES.get(self.culture_var.get(), {})
        is_garden = cfg.get("garden", False)

        # ── Параметри залежно від режиму ──────────────────────
        if is_garden:
            try:
                row_sp    = float(self._gv["row_sp"].get())
                plant_sp  = float(self._gv["plant_sp"].get())
                n_plot    = int(self._gv["plants_plot"].get())
                g_ends    = int(self._gv["guard_ends"].get())
                g_rows    = int(self._gv["guard_rows"].get())
            except ValueError:
                messagebox.showwarning("", "Перевірте параметри садіння."); return
            pw = row_sp; pl = plant_sp  # для сумісності
        else:
            try:
                pw = float(self._pv["pw"].get())
                pl = float(self._pv["pl"].get())
            except ValueError:
                messagebox.showwarning("", "Перевірте розміри ділянки."); return
            row_sp = pw; plant_sp = pl
            n_plot = 1; g_ends = 0; g_rows = 0

        design = self.design_var.get()
        rng = random.Random(seed)
        k = len(variants)

        if design == "latin" and k > 8:
            messagebox.showwarning("Латинський квадрат",
                f"Максимум 8 варіантів. У вас {k}."); return

        plan = []

        if design == "crd":
            all_p = variants * reps; rng.shuffle(all_p)
            for i, v in enumerate(all_p):
                plan.append({"plot": i+1, "rep": "–",
                             "variant": v, "row": i//k+1, "col": i%k+1})

        elif design == "rcbd":
            pn = 0
            for b in range(1, reps+1):
                bv = variants[:]; rng.shuffle(bv)
                for i, v in enumerate(bv):
                    pn += 1
                    plan.append({"plot": pn, "rep": f"Повт. {b}",
                                "variant": v, "row": b, "col": i+1})

        elif design == "latin":
            reps = k
            base = list(range(k))
            rows_p = [base[:]]
            for _ in range(k-1):
                rows_p.append(rows_p[-1][1:] + [rows_p[-1][0]])
            rng.shuffle(rows_p)
            cp = list(range(k)); rng.shuffle(cp)
            pn = 0
            for r in range(k):
                for c in range(k):
                    pn += 1
                    plan.append({"plot": pn, "rep": f"Рядок {r+1}",
                                "variant": variants[rows_p[r][cp[c]]],
                                "row": r+1, "col": c+1,
                                "col_label": f"Стовп. {c+1}"})

        elif design == "split":
            sp_vars = [v.strip() for v in
                       self.sp_text.get("1.0", "end").splitlines() if v.strip()]
            if len(sp_vars) < 2:
                messagebox.showwarning("", "Введіть щонайменше 2 sub-plot варіанти."); return
            pn = 0
            for b in range(1, reps+1):
                wp_o = variants[:]; rng.shuffle(wp_o)
                for wp in wp_o:
                    sp_o = sp_vars[:]; rng.shuffle(sp_o)
                    for sp in sp_o:
                        pn += 1
                        plan.append({"plot": pn, "rep": f"Повт. {b}",
                                    "variant": f"{wp} / {sp}",
                                    "wp": wp, "sp": sp, "row": b,
                                    "col": wp_o.index(wp)*len(sp_vars)+sp_o.index(sp)+1})

        design_name = {v: l for v, l, _ in self.DESIGNS}.get(design, design)

        # Розраховуємо площу
        if is_garden:
            plot_area = n_plot * row_sp * plant_sp
            total_area = plot_area * len(plan)
            area_msg = (f"Облікових рослин/ділянку: {n_plot}\n"
                       f"Схема садіння: {row_sp}×{plant_sp} м\n"
                       f"Площа ділянки: {plot_area:.1f} м²\n"
                       f"Загальна площа (облікова): {total_area:.0f} м²")
        else:
            plot_area = pw * pl
            total_area = plot_area * len(plan)
            area_msg = f"Площа ділянки: {pw}×{pl} м = {plot_area:.0f} м²\nЗагальна площа: {total_area:.0f} м²"

        self._plan_data = {
            "plan": plan, "variants": variants, "reps": reps,
            "design": design, "design_name": design_name,
            "seed": seed, "k": k, "pw": pw, "pl": pl,
            "is_garden": is_garden,
            "row_sp": row_sp if is_garden else pw,
            "plant_sp": plant_sp if is_garden else pl,
            "n_plot": n_plot, "g_ends": g_ends, "g_rows": g_rows,
            "culture": self.culture_var.get(),
            "unit": cfg.get("unit", "ділянка"),
            "name": self._nv["name"].get(),
            "year": self._nv["year"].get(),
            "loc":  self._nv["loc"].get(),
            "resp": self._nv["resp"].get(),
        }

        self._draw_scheme()
        self._fill_rand()
        self._fill_journal()
        self.nb.select(0)

        messagebox.showinfo("Готово",
            f"План згенеровано!\n\n"
            f"Дизайн: {design_name}\n"
            f"Варіантів: {k}   |   Повторностей: {reps}\n"
            f"Ділянок: {len(plan)}\n"
            f"{area_msg}\n\n"
            f"Seed рандомізації: {seed}\n"
            f"⚠ Збережіть seed у документацію!")
    # ═══════════════════════════════════════════════════════
    # Польова схема
    # ═══════════════════════════════════════════════════════
    def _draw_scheme(self):
        if not self._plan_data: return
        cv = self._scheme_cv; cv.delete("all")
        d = self._plan_data; plan = d["plan"]
        is_garden = d.get("is_garden", False)

        if is_garden:
            self._draw_scheme_garden(cv, d, plan)
        else:
            self._draw_scheme_field(cv, d, plan)

    def _draw_scheme_field(self, cv, d, plan):
        """Польова схема — кожна клітинка = ділянка."""
        PALETTES = ["#aed6f1","#a9dfbf","#f9e79f","#f1948a","#d2b4de",
                    "#a3e4d7","#fad7a0","#d5d8dc","#82e0aa","#f0b27a",
                    "#85c1e9","#f7dc6f","#c39bd3","#76d7c4","#f8c471"]
        all_v = list(dict.fromkeys(p["variant"] for p in plan))
        cmap  = {v: PALETTES[i % len(PALETTES)] for i, v in enumerate(all_v)}
        cols_set = sorted(set(p["col"] for p in plan))
        rows_set = sorted(set(p["row"] for p in plan))
        nc = len(cols_set); nr = len(rows_set)
        cw = max(78, min(130, 680 // nc))
        ch = max(44, min(72,  400 // nr))
        pad = 5; x0 = 110; y0 = 55

        title = d.get("name") or "План досліду"
        sub   = f"{d.get('culture','')}  |  {d['design_name']}  |  Seed={d['seed']}  |  {d.get('year','')}"
        cv.create_text(x0+nc*cw//2, 16, text=title,
                       font=("Times New Roman",12,"bold"), fill="#000")
        cv.create_text(x0+nc*cw//2, 34, text=sub,
                       font=("Times New Roman",9), fill="#555")

        rep_map = {}
        for p in plan: rep_map[p["row"]] = p["rep"]
        for i, r in enumerate(rows_set):
            cv.create_text(x0-6, y0+i*ch+ch//2,
                           text=rep_map.get(r, f"Ряд {r}"),
                           anchor="e", font=("Times New Roman",9,"bold"), fill="#333")
        for j, c in enumerate(cols_set):
            lbl = next((p for p in plan if p["col"]==c), {}).get("col_label", f"#{c}")
            cv.create_text(x0+j*cw+cw//2, y0-14,
                           text=lbl, font=("Times New Roman",8), fill="#555")
        for p in plan:
            ci = cols_set.index(p["col"]); ri = rows_set.index(p["row"])
            x1,y1 = x0+ci*cw, y0+ri*ch; x2,y2 = x1+cw-pad, y1+ch-pad
            cv.create_rectangle(x1,y1,x2,y2,
                                fill=cmap.get(p["variant"],"#eee"),
                                outline="#888", width=1)
            cv.create_text(x1+5,y1+6, text=f"№{p['plot']}",
                           anchor="nw", font=("Courier New",7), fill="#555")
            short = p["variant"][:14]+"…" if len(p["variant"])>14 else p["variant"]
            cv.create_text((x1+x2)//2,(y1+y2)//2, text=short,
                           font=("Times New Roman",8), fill="#000", width=cw-10)

        leg_y = y0+nr*ch+16
        cv.create_text(x0, leg_y, text="Легенда:",
                       anchor="w", font=("Times New Roman",10,"bold"))
        cpr = 3
        for i,v in enumerate(all_v):
            lx = x0+(i%cpr)*240; ly = leg_y+18+(i//cpr)*20
            cv.create_rectangle(lx,ly,lx+13,ly+13, fill=cmap[v], outline="#888")
            cv.create_text(lx+17,ly+7, text=v, anchor="w", font=("Times New Roman",9))

        tot_w = x0+nc*cw+20
        tot_h = leg_y+22*(len(all_v)//cpr+2)+10
        cv.configure(scrollregion=(0,0,tot_w,tot_h))

    def _draw_scheme_garden(self, cv, d, plan):
        """
        Садівнича схема — кожна клітинка = одна рослина.
        Захисні рослини на початку/кінці ряду — сірі.
        Захисні ряди між повторностями — штриховані.
        """
        PALETTES = ["#aed6f1","#a9dfbf","#f9e79f","#f1948a","#d2b4de",
                    "#a3e4d7","#fad7a0","#d5d8dc","#82e0aa","#f0b27a",
                    "#85c1e9","#f7dc6f","#c39bd3","#76d7c4","#f8c471"]
        GUARD_COLOR  = "#d0d0d0"   # захисні рослини
        GUARD_STIPPLE = "gray50"   # захисний ряд (штрих)

        all_v = list(dict.fromkeys(p["variant"] for p in plan))
        cmap  = {v: PALETTES[i % len(PALETTES)] for i, v in enumerate(all_v)}

        n_plot   = d.get("n_plot", 5)
        g_ends   = d.get("g_ends", 1)
        g_rows   = d.get("g_rows", 1)
        reps     = d.get("reps",   3)
        k        = d.get("k",      4)   # кількість варіантів
        design   = d.get("design", "rcbd")
        row_sp   = d.get("row_sp",  4.0)
        plant_sp = d.get("plant_sp",5.0)
        unit     = d.get("unit",   "дерево")

        # Загальна кількість рослин у ряду:
        # захисні_поч + [варіант_1 ... варіант_k] (повторяється n_plot) + захисні_кін
        # Для RCBD: 1 ряд = 1 повторність = k варіантів × n_plot рослин
        total_plants_row = g_ends + k * n_plot + g_ends
        # Рядів на схемі: reps повторностей + (reps-1)*g_rows захисних рядів
        total_rows = reps + (reps - 1) * g_rows

        # Розмір клітинки
        cw = max(28, min(60, 700 // total_plants_row))
        ch = max(28, min(55, 500 // total_rows))
        pad = 3; x0 = 120; y0 = 70

        # Заголовок
        title = d.get("name") or "Садівничий дослід"
        sub   = (f"{d.get('culture','')}  |  {d['design_name']}  |  "
                 f"Схема: {row_sp}×{plant_sp} м  |  "
                 f"Рослин/ділянку: {n_plot}  |  Seed={d['seed']}")
        cv.create_text(x0 + total_plants_row*cw//2, 18,
                       text=title, font=("Times New Roman",12,"bold"), fill="#000")
        cv.create_text(x0 + total_plants_row*cw//2, 36,
                       text=sub, font=("Times New Roman",9), fill="#555")

        # Мітки колонок (позиції рослин)
        for ci in range(total_plants_row):
            if ci < g_ends or ci >= total_plants_row - g_ends:
                lbl = "З"   # Захисна
            else:
                pos = ci - g_ends
                var_idx = pos // n_plot
                plant_in_var = pos % n_plot + 1
                lbl = f"#{plant_in_var}"
            cv.create_text(x0+ci*cw+cw//2, y0-14,
                           text=lbl, font=("Times New Roman",7), fill="#777")

        # Рядки (повторності + захисні ряди)
        row_screen = 0   # лічильник рядків на екрані
        for rep_idx in range(reps):
            # Мітка повторності
            row_y = y0 + row_screen * ch
            cv.create_text(x0-6, row_y+ch//2,
                           text=f"Повт.{rep_idx+1}",
                           anchor="e", font=("Times New Roman",8,"bold"), fill="#333")

            # Беремо план для цієї повторності
            rep_plan = [p for p in plan if p["row"] == rep_idx+1]
            # Будуємо порядок варіантів у ряду
            col_to_var = {p["col"]: p["variant"] for p in rep_plan}

            # Малюємо рослини в ряду
            for ci in range(total_plants_row):
                x1 = x0 + ci * cw; y1 = row_y
                x2 = x1 + cw - pad; y2 = y1 + ch - pad

                if ci < g_ends or ci >= total_plants_row - g_ends:
                    # Захисна рослина на початку/кінці
                    cv.create_oval(x1+2, y1+2, x2-2, y2-2,
                                   fill=GUARD_COLOR, outline="#999", width=1)
                    cv.create_text((x1+x2)//2, (y1+y2)//2,
                                   text="З", font=("Times New Roman",7), fill="#888")
                else:
                    pos = ci - g_ends
                    var_col = pos // n_plot + 1  # номер варіанту (стовпець)
                    variant = col_to_var.get(var_col, "")
                    color = cmap.get(variant, "#eee")
                    # Коло = дерево/кущ
                    cv.create_oval(x1+2, y1+2, x2-2, y2-2,
                                   fill=color, outline="#666", width=1)
                    plant_in_var = pos % n_plot + 1
                    cv.create_text((x1+x2)//2, (y1+y2)//2,
                                   text=str(plant_in_var),
                                   font=("Times New Roman",7), fill="#000")

            row_screen += 1

            # Захисні ряди між повторностями
            if rep_idx < reps - 1:
                for gr in range(g_rows):
                    gy = y0 + row_screen * ch
                    cv.create_text(x0-6, gy+ch//2,
                                   text="Захисний", anchor="e",
                                   font=("Times New Roman",7,"italic"), fill="#aaa")
                    for ci in range(total_plants_row):
                        x1 = x0+ci*cw; y1 = gy
                        x2 = x1+cw-pad; y2 = y1+ch-pad
                        cv.create_oval(x1+2, y1+2, x2-2, y2-2,
                                       fill="#e8e8e8", outline="#bbb",
                                       width=1, stipple=GUARD_STIPPLE
                                       if GUARD_STIPPLE else "")
                    row_screen += 1

        # Роздільники між варіантами у рядку (вертикальні лінії)
        for var_i in range(k+1):
            lx = x0 + (g_ends + var_i * n_plot) * cw
            cv.create_line(lx, y0-5, lx, y0+reps*ch + (reps-1)*g_rows*ch + 5,
                           fill="#1a4b8c", width=1, dash=(4,3))

        # Легенда варіантів
        leg_y = y0 + total_rows * ch + 20
        cv.create_text(x0, leg_y, text="Легенда (варіанти):",
                       anchor="w", font=("Times New Roman",10,"bold"))
        cpr = 3
        for i, v in enumerate(all_v):
            lx = x0 + (i%cpr)*260; ly = leg_y+18+(i//cpr)*22
            cv.create_oval(lx, ly, lx+14, ly+14, fill=cmap[v], outline="#666")
            cv.create_text(lx+18, ly+7, text=v, anchor="w",
                           font=("Times New Roman",9))

        leg_y2 = leg_y + 18 + ((len(all_v)-1)//cpr+1)*22 + 8
        # Пояснення символів
        cv.create_oval(x0, leg_y2, x0+14, leg_y2+14,
                       fill=GUARD_COLOR, outline="#999")
        cv.create_text(x0+18, leg_y2+7,
                       text="З — захисна рослина (не обліковується)",
                       anchor="w", font=("Times New Roman",9), fill="#666")
        cv.create_oval(x0, leg_y2+20, x0+14, leg_y2+34,
                       fill="#e8e8e8", outline="#bbb")
        cv.create_text(x0+18, leg_y2+27,
                       text="Захисний ряд між повторностями",
                       anchor="w", font=("Times New Roman",9), fill="#666")
        # Схема садіння
        cv.create_text(x0, leg_y2+50,
                       text=f"Схема садіння: {row_sp} м × {plant_sp} м  |  "
                            f"Облікових {unit}/ділянку: {n_plot}  |  "
                            f"Захисних з кожного боку: {g_ends}",
                       anchor="w", font=("Times New Roman",9), fill="#333")

        tot_w = x0 + total_plants_row*cw + 20
        tot_h = leg_y2 + 70
        cv.configure(scrollregion=(0, 0, tot_w, tot_h))

    # ═══════════════════════════════════════════════════════
    # Список рандомізації
    # ═══════════════════════════════════════════════════════
    def _fill_rand(self):
        if not self._plan_data: return
        d = self._plan_data; plan = d["plan"]
        self.rand_txt.configure(state="normal")
        self.rand_txt.delete("1.0", tk.END)
        lines = [
            "═"*62,
            "     СПИСОК РАНДОМІЗАЦІЇ ПОЛЬОВОГО ДОСЛІДУ",
            "═"*62,
            f"  Назва:          {d.get('name') or '—'}",
            f"  Рік:            {d.get('year','')}",
            f"  Місце:          {d.get('loc') or '—'}",
            f"  Відповідальний: {d.get('resp') or '—'}",
            f"  Культура:       {d.get('culture','')}",
            f"  Дизайн:         {d['design_name']}",
            f"  Варіантів:      {d['k']}",
            f"  Повторностей:   {d['reps']}",
            f"  Ділянок:        {len(plan)}",
            f"  Площа ділянки:  {d['pw']} × {d['pl']} м = {d['pw']*d['pl']:.1f} м²",
            f"  Загальна площа: {d['pw']*d['pl']*len(plan):.0f} м²",
            f"  Seed рандом.:   {d['seed']}  ← зберігайте цей номер!",
            "─"*62,
            f"  {'№':<6}  {'Повторність':<16}  Варіант",
            "─"*62,
        ]
        for p in sorted(plan, key=lambda x: x["plot"]):
            lines.append(f"  {p['plot']:<6}  {p['rep']:<16}  {p['variant']}")
        lines += [
            "─"*62,
            f"  Сформовано: {datetime.now().strftime('%d.%m.%Y  %H:%M')}",
        ]
        self.rand_txt.insert("1.0", "\n".join(lines))
        self.rand_txt.configure(state="disabled")

    # ═══════════════════════════════════════════════════════
    # Польовий журнал
    # ═══════════════════════════════════════════════════════
    def _fill_journal(self):
        if not self._plan_data: return
        d    = self._plan_data
        plan = d["plan"]
        is_garden = d.get("is_garden", False)
        n_plot    = d.get("n_plot", 1) if is_garden else 1
        unit      = d.get("unit", "ділянка")

        ind_text   = self.ind_var.get().strip()
        indicators = [s.strip() for s in ind_text.split(";") if s.strip()]
        if not indicators:
            indicators = ["Показник 1","Показник 2","Показник 3","Показник 4"]

        # Зберігаємо поточні назви заголовків для перейменування
        if not hasattr(self, '_ind_vars'):
            self._ind_vars = {}
        self._cur_indicators = indicators[:]

        for item in self.journal_tv.get_children():
            self.journal_tv.delete(item)

        # Стиль Treeview — чіткі межі
        style = ttk.Style()
        style.configure("Journal.Treeview",
                        rowheight=22,
                        font=("Times New Roman",10),
                        borderwidth=1)
        style.configure("Journal.Treeview.Heading",
                        font=("Times New Roman",10,"bold"),
                        background="#1a4b8c",
                        foreground="white",
                        relief="flat")
        style.map("Journal.Treeview.Heading",
                  background=[("active","#2a5ba8")])
        self.journal_tv.configure(style="Journal.Treeview")

        if is_garden:
            cols = ("№ діл.", "Повт.", f"{unit.capitalize()} №",
                    "Варіант") + tuple(indicators) + ("Примітки",)
        else:
            cols = ("№ діл.", "Повторність",
                    "Варіант") + tuple(indicators) + ("Примітки",)

        self.journal_tv["columns"] = cols
        self.journal_tv["show"]    = "headings"

        w_map = {"№ діл.":55, "Повт.":70, "Повторність":110,
                 f"{unit.capitalize()} №":70, "Варіант":160, "Примітки":100}
        for col in cols:
            self.journal_tv.heading(col, text=col)
            w = w_map.get(col, 100)
            self.journal_tv.column(col, width=w, minwidth=40,
                                   anchor="center" if w<130 else "w",
                                   stretch=True)

        # Подвійний клік на заголовок — перейменування показника
        def _on_heading_dbl(event):
            region = self.journal_tv.identify_region(event.x, event.y)
            if region != "heading": return
            col_id = self.journal_tv.identify_column(event.x)
            col_n  = int(col_id[1:]) - 1  # 0-based
            cur_name = cols[col_n]
            # Лише показники (не службові колонки)
            fixed = (3 if is_garden else 2) + 1  # перші fixed колонок - службові
            if col_n < fixed: return

            ind_idx = col_n - fixed
            if ind_idx >= len(indicators): return

            dlg = tk.Toplevel(self.win); dlg.title("Перейменувати показник")
            dlg.resizable(False,False); dlg.grab_set(); set_icon(dlg)
            tk.Label(dlg, text="Нова назва показника:",
                     font=("Times New Roman",12)).pack(padx=16, pady=(14,4))
            tv2 = tk.StringVar(value=cur_name)
            e = tk.Entry(dlg, textvariable=tv2,
                         font=("Times New Roman",12), width=28)
            e.pack(padx=16, pady=4)
            e.select_range(0, tk.END); e.focus_set()
            def _ok():
                nm = tv2.get().strip()
                if nm:
                    indicators[ind_idx] = nm
                    # Оновити поле indicators
                    self.ind_var.set("; ".join(indicators))
                    dlg.destroy()
                    self._fill_journal()
            tk.Button(dlg, text="ОК", bg="#c62828", fg="white",
                      font=("Times New Roman",12), command=_ok
                      ).pack(pady=(4,14))
            dlg.bind("<Return>", lambda e2: _ok())
            center_win(dlg)

        self.journal_tv.bind("<Double-1>", _on_heading_dbl)

        # Теги для чергування рядків
        self.journal_tv.tag_configure("even", background="#f0f4ff")
        self.journal_tv.tag_configure("odd",  background="#ffffff")

        row_i = 0
        for p in sorted(plan, key=lambda x: x["plot"]):
            if is_garden:
                # Кожна облікова рослина — окремий рядок
                for plant_n in range(1, n_plot + 1):
                    tag = "even" if row_i % 2 == 0 else "odd"
                    vals = (p["plot"], p["rep"], plant_n,
                            p["variant"]) + tuple("" for _ in indicators) + ("",)
                    self.journal_tv.insert("", "end", values=vals, tags=(tag,))
                    row_i += 1
            else:
                tag = "even" if row_i % 2 == 0 else "odd"
                vals = (p["plot"], p["rep"],
                        p["variant"]) + tuple("" for _ in indicators) + ("",)
                self.journal_tv.insert("", "end", values=vals, tags=(tag,))
                row_i += 1

    def _refresh_journal(self):
        if not self._plan_data:
            messagebox.showwarning("","Спочатку згенеруйте план."); return
        self._fill_journal()

    # ═══════════════════════════════════════════════════════
    # Збереження
    # ═══════════════════════════════════════════════════════
    def _save_rand_txt(self):
        if not self._plan_data:
            messagebox.showwarning("","Спочатку згенеруйте план."); return
        path = filedialog.asksaveasfilename(
            parent=self.win, defaultextension=".txt",
            filetypes=[("Текстовий файл","*.txt")],
            title="Зберегти список рандомізації")
        if not path: return
        try:
            with open(path,"w",encoding="utf-8") as f:
                f.write(self.rand_txt.get("1.0",tk.END))
            messagebox.showinfo("Збережено",f"Збережено:\n{path}")
        except Exception as ex:
            messagebox.showerror("Помилка",str(ex))

    def _save_excel(self):
        if not self._plan_data:
            messagebox.showwarning("","Спочатку згенеруйте план."); return
        if not HAS_OPENPYXL:
            messagebox.showerror("","Потрібен openpyxl: pip install openpyxl"); return

        path = filedialog.asksaveasfilename(
            parent=self.win, defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],
            title="Зберегти польовий журнал")
        if not path: return

        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            d  = self._plan_data; plan = d["plan"]
            ind_text   = self.ind_var.get().strip()
            indicators = [s.strip() for s in ind_text.split(";") if s.strip()]
            if not indicators:
                indicators = ["Показник 1","Показник 2","Показник 3","Показник 4"]

            wb  = openpyxl.Workbook()
            hfill = PatternFill("solid", fgColor="1A4B8C")
            hfont = Font(color="FFFFFF", bold=True,
                         name="Times New Roman", size=11)
            nfont = Font(name="Times New Roman", size=11)
            bfont = Font(name="Times New Roman", size=11, bold=True)
            ca    = Alignment(horizontal="center", vertical="center",
                              wrap_text=True)
            thin  = Side(style="thin", color="AAAAAA")
            brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

            is_garden = d.get("is_garden", False)
            n_plot    = d.get("n_plot", 1) if is_garden else 1
            unit      = d.get("unit", "ділянка")

            # ── Лист 1: Польовий журнал ─────────────────────
            ws = wb.active; ws.title = "Польовий журнал"

            # Шапка
            if is_garden:
                info_rows = [
                    "ПОЛЬОВИЙ ЖУРНАЛ СПОСТЕРЕЖЕНЬ",
                    f"Дослід: {d.get('name') or '—'}",
                    f"Рік: {d.get('year','')}    Місце: {d.get('loc') or '—'}    Відповідальний: {d.get('resp') or '—'}",
                    f"Культура: {d.get('culture','')}    Дизайн: {d['design_name']}",
                    f"Варіантів: {d['k']}    Повторностей: {d['reps']}    "
                    f"Облікових {unit}/ділянку: {n_plot}    "
                    f"Схема: {d.get('row_sp','?')}×{d.get('plant_sp','?')} м    "
                    f"Захисних: {d.get('g_ends',0)} з кожного боку",
                ]
            else:
                info_rows = [
                    "ПОЛЬОВИЙ ЖУРНАЛ СПОСТЕРЕЖЕНЬ",
                    f"Дослід: {d.get('name') or '—'}",
                    f"Рік: {d.get('year','')}    Місце: {d.get('loc') or '—'}    Відповідальний: {d.get('resp') or '—'}",
                    f"Культура: {d.get('culture','')}    Дизайн: {d['design_name']}    "
                    f"Варіантів: {d['k']}    Повторностей: {d['reps']}    "
                    f"Ділянок: {len(plan)}    Площа: {d['pw']}×{d['pl']} м",
                ]
            for ri, txt in enumerate(info_rows, 1):
                c = ws.cell(ri, 1, txt)
                c.font = bfont if ri == 1 else nfont

            # Таблиця журналу
            hr = len(info_rows) + 2
            if is_garden:
                j_hdrs = ["№ діл.", "Повт.", f"{unit.capitalize()} №", "Варіант"] + indicators + ["Примітки"]
            else:
                j_hdrs = ["№ ділянки","Повторність","Варіант"] + indicators + ["Примітки"]

            for ci, h in enumerate(j_hdrs, 1):
                c = ws.cell(hr, ci, h)
                c.fill = hfill; c.font = hfont
                c.alignment = ca; c.border = brd

            PALETTES_HEX = ["AED6F1","A9DFBF","F9E79F","F1948A","D2B4DE",
                            "A3E4D7","FAD7A0","D5D8DC","82E0AA","F0B27A"]
            all_v = list(dict.fromkeys(p["variant"] for p in plan))
            vcols = {v: PALETTES_HEX[i % len(PALETTES_HEX)]
                     for i, v in enumerate(all_v)}

            row_excel = hr + 1
            row_i = 0
            for p in sorted(plan, key=lambda x: x["plot"]):
                if is_garden:
                    for plant_n in range(1, n_plot + 1):
                        even = row_i % 2 == 0
                        rfill = PatternFill("solid", fgColor="EEF4FF" if even else "FFFFFF")
                        vals = [p["plot"], p["rep"], plant_n, p["variant"]] + [""] * len(indicators) + [""]
                        for ci, val in enumerate(vals, 1):
                            c = ws.cell(row_excel, ci, val)
                            c.font = nfont; c.alignment = ca; c.border = brd
                            if ci <= 4 and even: c.fill = rfill
                        row_excel += 1; row_i += 1
                else:
                    even = row_i % 2 == 0
                    rfill = PatternFill("solid", fgColor="EEF4FF" if even else "FFFFFF")
                    vals = [p["plot"], p["rep"], p["variant"]] + [""] * len(indicators) + [""]
                    for ci, val in enumerate(vals, 1):
                        c = ws.cell(row_excel, ci, val)
                        c.font = nfont; c.alignment = ca; c.border = brd
                        if ci <= 3 and even: c.fill = rfill
                    row_excel += 1; row_i += 1

            # Ширини стовпців
            w_cols = ([8, 10, 10, 32] if is_garden else [9, 14, 32]) + [14]*len(indicators) + [16]
            for ci, w in enumerate(w_cols, 1):
                if ci <= 26:
                    ws.column_dimensions[chr(64+ci)].width = w
            ws.row_dimensions[hr].height = 30

            # ── Лист 2: Рандомізація ────────────────────────
            ws2 = wb.create_sheet("Рандомізація")
            r_hdrs = ["№ ділянки","Повторність","Варіант"]
            for ci, h in enumerate(r_hdrs, 1):
                c = ws2.cell(1, ci, h)
                c.fill = hfill; c.font = hfont
                c.alignment = ca; c.border = brd
            for ri, p in enumerate(sorted(plan, key=lambda x: x["plot"])):
                row = 2 + ri
                fc = PatternFill("solid",
                                 fgColor=vcols.get(p["variant"],"EEEEEE"))
                for ci, val in enumerate(
                    [p["plot"], p["rep"], p["variant"]], 1
                ):
                    c = ws2.cell(row, ci, val)
                    c.font = nfont; c.alignment = ca; c.border = brd
                    if ci == 3: c.fill = fc
            for ci, w in zip([1,2,3],[9,14,36]):
                ws2.column_dimensions[chr(64+ci)].width = w

            wb.save(path)
            messagebox.showinfo("Збережено",
                f"Збережено:\n{path}\n\n"
                "Лист 1 — Польовий журнал\n"
                "Лист 2 — Рандомізація")
        except Exception as ex:
            messagebox.showerror("Помилка збереження", str(ex))

    def _save_png(self):
        if not self._plan_data:
            messagebox.showwarning("","Спочатку згенеруйте план."); return
        if not HAS_MPL:
            messagebox.showwarning("","matplotlib недоступний."); return

        path = filedialog.asksaveasfilename(
            parent=self.win, defaultextension=".png",
            filetypes=[("PNG зображення","*.png")],
            title="Зберегти схему як PNG")
        if not path: return

        d  = self._plan_data; plan = d["plan"]
        all_v = list(dict.fromkeys(p["variant"] for p in plan))
        PALETTES = ["#aed6f1","#a9dfbf","#f9e79f","#f1948a","#d2b4de",
                    "#a3e4d7","#fad7a0","#d5d8dc","#82e0aa","#f0b27a"]
        cmap = {v: PALETTES[i%len(PALETTES)] for i,v in enumerate(all_v)}

        cols_set = sorted(set(p["col"] for p in plan))
        rows_set = sorted(set(p["row"] for p in plan))
        nc = len(cols_set); nr = len(rows_set)

        fig = Figure(figsize=(max(10, nc*1.6+2), max(5, nr*1.1+3)), dpi=130)
        ax  = fig.add_axes([0.09, 0.16, 0.88, 0.72])
        ax.set_xlim(0, nc); ax.set_ylim(0, nr); ax.set_aspect("equal")
        ax.axis("off")

        rep_map = {}
        for p in plan: rep_map[rows_set.index(p["row"])] = p["rep"]

        # Ділянки
        for p in plan:
            ci = cols_set.index(p["col"])
            ri = rows_set.index(p["row"])
            rect = matplotlib.patches.FancyBboxPatch(
                (ci+0.04, nr-ri-0.95), 0.91, 0.89,
                boxstyle="round,pad=0.02",
                facecolor=cmap.get(p["variant"],"#eee"),
                edgecolor="#777", linewidth=0.7)
            ax.add_patch(rect)
            short = (p["variant"][:13]+"…"
                     if len(p["variant"]) > 13 else p["variant"])
            ax.text(ci+0.5, nr-ri-0.5, short, ha="center", va="center",
                    fontsize=6.5, fontfamily="Times New Roman")
            ax.text(ci+0.07, nr-ri-0.1, f"#{p['plot']}",
                    ha="left", va="top", fontsize=5.5, color="#555",
                    fontfamily="Courier New")

        # Мітки рядків і стовпців
        for i,r in enumerate(rows_set):
            ax.text(-0.08, nr-i-0.5, rep_map.get(i,""),
                    ha="right", va="center", fontsize=7,
                    fontfamily="Times New Roman")
        for j,c in enumerate(cols_set):
            p_ = next(p for p in plan if p["col"]==c)
            lbl = p_.get("col_label", f"#{c}")
            ax.text(j+0.5, nr+0.08, lbl, ha="center", va="bottom",
                    fontsize=7, fontfamily="Times New Roman")

        # Заголовок
        name = d.get("name") or "План досліду"
        fig.suptitle(
            f"{name}  |  {d.get('year','')}",
            fontsize=11, fontfamily="Times New Roman", fontweight="bold", y=0.98)
        ax.set_title(
            f"Культура: {d.get('culture','')}  |  Дизайн: {d['design_name']}  "
            f"|  Варіантів: {d['k']}  |  Повторностей: {d['reps']}  "
            f"|  Seed: {d['seed']}",
            fontsize=8, fontfamily="Times New Roman")

        # Легенда
        from matplotlib.patches import Patch
        handles = [Patch(facecolor=cmap[v], edgecolor="#777", label=v)
                   for v in all_v]
        fig.legend(handles=handles, loc="lower center",
                   ncol=min(4, len(all_v)), fontsize=7,
                   framealpha=0.8, bbox_to_anchor=(0.5, 0.01))
        try:
            fig.savefig(path, dpi=150, bbox_inches="tight")
            messagebox.showinfo("Збережено", f"PNG збережено:\n{path}")
            open_file_cross(path)
        except Exception as ex:
            messagebox.showerror("Помилка", str(ex))
